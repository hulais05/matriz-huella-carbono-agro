"""
Dashboard de Huella de Carbono Corporativa — Sector Agropecuario
Lee directamente la Matriz Excel y se actualiza al detectar cambios en el archivo.

Lanzar con:  streamlit run dashboard_carbono.py
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import openpyxl
import os
import time
from datetime import datetime

# ─── CONFIG ───────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Huella de Carbono | Agropecuario",
    page_icon="🌿",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── TEMA (oscuro / claro / según sistema) ────────────────────────────────────
# st.context.theme.type refleja el tema realmente activo (incluye "Use system
# setting" ya resuelto), así los gráficos y tarjetas custom se adaptan también
# cuando el usuario cambia de tema desde el menú ⋮ de Streamlit.
DARK = (st.context.theme.type or "dark") != "light"

if DARK:
    T = dict(
        bg_app="#0E1117", bg_card="#161B27", bg_alt="#1A1F2E",
        text="#FAFAFA", text_soft="#CCCCCC", text_muted="#888888", text_dim="#666666",
        grid="#333333", border="rgba(255,255,255,0.08)",
        card_grad=("rgba(255,255,255,0.04)", "rgba(255,255,255,0.01)"),
        h3="#4FC3F7",
        gauge_steps=["#1A0A0A", "#1A1A0A", "#0F1F0A", "#0A1A0A"],
    )
else:
    T = dict(
        bg_app="#FFFFFF", bg_card="#F0F2F6", bg_alt="#E6E9EF",
        text="#0E1117", text_soft="#333333", text_muted="#666666", text_dim="#999999",
        grid="#DDDDDD", border="rgba(0,0,0,0.08)",
        card_grad=("rgba(0,0,0,0.03)", "rgba(0,0,0,0.01)"),
        h3="#0070C0",
        gauge_steps=["#FBE7E7", "#FBF3DC", "#E6F4E6", "#E0F0DC"],
    )

# En uso local apunta al archivo de trabajo en Downloads. Si no existe (p.ej. en
# Streamlit Cloud, donde no hay carpeta Downloads), usa la copia incluida en el
# repo junto a este script para que la demo funcione igual.
_LOCAL_EXCEL   = os.path.expanduser("~/Downloads/Matriz_HuellaCarbono_Agropecuario.xlsx")
_BUNDLED_EXCEL = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Matriz_HuellaCarbono_Agropecuario.xlsx")
EXCEL_DEFAULT  = _LOCAL_EXCEL if os.path.exists(_LOCAL_EXCEL) else _BUNDLED_EXCEL

PRG_CH4_FOSIL = 29.8   # AR6 — CH4 fósil (combustión gasoil/GLP)
PRG_CH4_BIO   = 27.2   # AR6 — CH4 biogénico (ganadería/efluentes)
PRG_N2O = 273.0
FE_GASOIL_MOV  = 2.701    # kg CO2/L — combustión móvil
FE_GASOIL_EST  = 2.701    # kg CO2/L — combustión estacionaria
FE_GLP         = 2.983    # kg CO2/kg GLP
FE_RED_ARG     = 0.383    # kg CO2e/kWh — CAMMESA 2023 (verificar valor vigente cada año en cammesaweb.cammesa.com)

CAT_LABELS = {
    "maquinaria":   "Combustión Móvil",
    "energia_est":  "Energía Estacionaria",
    "fermentacion": "Fermentación Entérica",
    "estiercol":    "Gestión de Estiércol",
    "suelos":       "Suelos Agrícolas",
    "efluentes":    "Efluentes",
    "electricidad": "Electricidad de Red (S2)",
}

SCOPE1_KEYS = ["maquinaria", "energia_est", "fermentacion", "estiercol", "suelos", "efluentes"]
COLORS_PIE  = ["#C00000", "#E05050", "#FF6B6B", "#FF9E9E", "#FF8000", "#FFA500", "#0070C0"]
COLOR_S2    = "#0070C0"
COLOR_SK    = "#00B050"
COLOR_NET   = "#4FC3F7"
COLOR_WARN  = "#ED7D31"

MESES = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]

# ─── HELPERS ──────────────────────────────────────────────────────────────────

def flt(val):
    """Convierte cualquier valor a float de forma segura."""
    if val is None or val == "" or str(val).strip() in ("---", "-", "N/A"):
        return 0.0
    try:
        return float(str(val).replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return 0.0

def fmt_t(val, decimals=1):
    if val == 0:
        return "—"
    return f"{val:,.{decimals}f}"

def pct_of(part, total):
    return (part / total * 100) if total and total != 0 else 0.0

# ─── LECTURA DE DATOS DESDE HOJAS FUENTE ─────────────────────────────────────

def read_maquinaria(wb):
    """S1-1 MAQUINARIA — rows 5-13, cols: B=nombre, E=meses, F=horas/mes, G=L/hora"""
    ws = wb["S1-1 MAQUINARIA"]
    rows, total = [], 0.0
    for r in range(5, 14):
        nombre  = ws.cell(r, 2).value or f"Máquina {r-4}"
        meses   = flt(ws.cell(r, 5).value)
        horas   = flt(ws.cell(r, 6).value)
        lhora   = flt(ws.cell(r, 7).value)
        consumo = meses * horas * lhora
        co2     = consumo * FE_GASOIL_MOV
        ch4     = consumo * 0.10 / 1000 * PRG_CH4_FOSIL
        n2o     = consumo * 0.60 / 1000 * PRG_N2O
        t       = (co2 + ch4 + n2o) / 1000
        rows.append({"nombre": str(nombre), "consumo_l": consumo, "tco2e": t})
        total += t
    return rows, total


def read_energia_estacionaria(wb):
    """S1-2 ENERGIA ESTAC — generadores rows 5-6, zeppelin row 11"""
    ws = wb["S1-2 ENERGIA ESTAC"]
    gen_l = sum(
        flt(ws.cell(r, 4).value) * flt(ws.cell(r, 5).value) * flt(ws.cell(r, 6).value)
        for r in range(5, 7)
    )
    co2_gen = gen_l * FE_GASOIL_EST / 1000

    glp_kg  = flt(ws.cell(11, 4).value) * flt(ws.cell(11, 5).value)
    co2_glp = glp_kg * FE_GLP / 1000

    return {
        "generadores_l":    gen_l,
        "co2_gen_tco2e":    co2_gen,
        "glp_kg":           glp_kg,
        "co2_glp_tco2e":    co2_glp,
        "total_tco2e":      co2_gen + co2_glp,
    }


def read_ganaderia(wb):
    """S1-3 GANADERIA — fermentación rows 6-15, estiércol rows 20-23"""
    ws = wb["S1-3 GANADERIA"]

    FE_FERM_DEF = [119.8, 100.0, 56.5, 47.8, 47.8, 27.0, 73.0, 8.0, 1.5, 0.02]
    FE_CH4_EST  = [16.0,  1.0,   1.0,  1.0]
    FE_N2O_EST  = [0.10,  0.20,  0.20, 0.07]
    CAT_FERM    = ["Vacas ordeñe","Vaquillonas tambo","Novillos feedlot","Novillitos feedlot",
                   "Vaquillonas recría","Terneros","Toros","Ovinos","Cerdos","Aves"]

    ferm_rows, ferm_total = [], 0.0
    for i, r in enumerate(range(6, 16)):
        cabs = flt(ws.cell(r, 3).value)
        dias = flt(ws.cell(r, 4).value)
        fe   = flt(ws.cell(r, 7).value) or FE_FERM_DEF[i]
        eq   = cabs * (dias / 365) if dias > 0 else cabs
        t    = eq * fe * PRG_CH4_BIO / 1000
        if cabs > 0:
            ferm_rows.append({"categoria": CAT_FERM[i], "cabezas": int(cabs), "tco2e": t})
        ferm_total += t

    est_total = 0.0
    for i, r in enumerate(range(20, 24)):
        cabs    = flt(ws.cell(r, 3).value)
        fe_ch4  = flt(ws.cell(r, 7).value) or FE_CH4_EST[i]
        fe_n2o  = flt(ws.cell(r, 8).value) or FE_N2O_EST[i]
        est_total += (cabs * fe_ch4 * PRG_CH4_BIO + cabs * fe_n2o * PRG_N2O) / 1000

    return {
        "fermentacion_tco2e": ferm_total,
        "estiercol_tco2e":    est_total,
        "ferm_detail":        ferm_rows,
    }


def read_suelos(wb):
    """S1-4 SUELOS AGRICOLAS — rows 5-13"""
    ws  = wb["S1-4 SUELOS AGRICOLAS"]
    CULT = ["Soja 1ra","Soja 2da","Maíz","Trigo","Sorgo","Girasol","Pasturas","Otro 1","Otro 2"]
    n2o_total, co2_urea_total, rows = 0.0, 0.0, []

    for i, r in enumerate(range(5, 14)):
        ha        = flt(ws.cell(r, 3).value)
        kg_n_ha   = flt(ws.cell(r, 4).value)
        kg_urea_ha= flt(ws.cell(r, 6).value)
        t_resid_ha= flt(ws.cell(r, 8).value)

        n_fert    = ha * kg_n_ha
        n_resid   = ha * t_resid_ha * 6.0         # 0.6% N en residuos → g/kg *1000
        n2o_kg    = (n_fert + n_resid) * 0.01 * (44/28)
        n2o_co2e  = n2o_kg * PRG_N2O / 1000
        co2_urea  = ha * kg_urea_ha * 0.734 / 1000
        t         = n2o_co2e + co2_urea

        if ha > 0 or kg_n_ha > 0:
            rows.append({"cultivo": CULT[i], "ha": ha, "tco2e": t})
        n2o_total      += n2o_co2e
        co2_urea_total += co2_urea

    return {
        "n2o_tco2e":     n2o_total,
        "co2_urea_tco2e":co2_urea_total,
        "total_tco2e":   n2o_total + co2_urea_total,
        "detail":        rows,
    }


def read_efluentes(wb):
    """S1-5 EFLUENTES — rows 5-8"""
    ws, total = wb["S1-5 EFLUENTES"], 0.0
    for r in range(5, 9):
        caudal  = flt(ws.cell(r, 3).value)
        dias    = flt(ws.cell(r, 4).value)
        dbo_in  = flt(ws.cell(r, 6).value)
        dbo_out = flt(ws.cell(r, 7).value)
        bo      = flt(ws.cell(r, 8).value) or 0.16
        mcf     = flt(ws.cell(r, 9).value) or 0.8
        vol     = caudal * dias
        ch4_m3  = vol * max(dbo_in - dbo_out, 0) / 1000 * bo * mcf
        total  += ch4_m3 * 0.00067 * 1000 * PRG_CH4_BIO / 1000
    return total


def read_electricidad(wb):
    """S2-1 ELECTRICIDAD — rows 5-16 (12 meses)"""
    ws, monthly, total = wb["S2-1 ELECTRICIDAD"], [], 0.0
    for i, r in enumerate(range(5, 17)):
        kwh   = flt(ws.cell(r, 3).value)
        costo = flt(ws.cell(r, 5).value)
        t     = kwh * FE_RED_ARG / 1000
        monthly.append({"mes": MESES[i], "kwh": kwh, "costo": costo, "tco2e": t})
        total += t
    return monthly, total


def read_sumideros(wb):
    """SUMIDEROS — monte rows 5-8, pasturas rows 13-16"""
    ws = wb["SUMIDEROS"]

    monte_rows, monte_total = [], 0.0
    for r in range(5, 9):
        ha      = flt(ws.cell(r, 3).value)
        fe      = flt(ws.cell(r, 7).value) or 3.2
        nombre  = ws.cell(r, 1).value or f"Unidad {r-4}"
        captura = -(ha * fe)
        if ha > 0:
            monte_rows.append({"nombre": str(nombre)[:35], "ha": ha, "tco2e": captura})
        monte_total += captura

    past_total = 0.0
    for r in range(13, 17):
        ha = flt(ws.cell(r, 3).value)
        fe = flt(ws.cell(r, 7).value) or 0.5
        past_total += -(ha * fe)

    return {
        "monte_tco2e":    monte_total,
        "pasturas_tco2e": past_total,
        "total_tco2e":    monte_total + past_total,
        "monte_detail":   monte_rows,
    }


def read_portada(wb):
    """PORTADA — datos del establecimiento."""
    ws = wb["PORTADA"]
    return {
        "nombre":    ws.cell(5,  2).value or "—",
        "provincia": ws.cell(7,  2).value or "—",
        "anio":      ws.cell(11, 2).value or "—",
        "ha_total":  ws.cell(9,  2).value or 1000,
        "ha_prod":   ws.cell(10, 2).value or 900,
    }


@st.cache_data(ttl=8)
def load_all(_mtime, path):
    """Lee y calcula todo. Cache invalidado cada 8 s o cuando mtime cambia."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        return None, str(e)

    try:
        portada  = read_portada(wb)
        maq_det, maq_t   = read_maquinaria(wb)
        energia          = read_energia_estacionaria(wb)
        gan              = read_ganaderia(wb)
        suelos           = read_suelos(wb)
        efluentes_t      = read_efluentes(wb)
        monthly_elec, e_t= read_electricidad(wb)
        sinks            = read_sumideros(wb)

        emissions = {
            "maquinaria":   maq_t,
            "energia_est":  energia["total_tco2e"],
            "fermentacion": gan["fermentacion_tco2e"],
            "estiercol":    gan["estiercol_tco2e"],
            "suelos":       suelos["total_tco2e"],
            "efluentes":    efluentes_t,
            "electricidad": e_t,
        }
        scope1  = sum(emissions[k] for k in SCOPE1_KEYS)
        scope2  = emissions["electricidad"]
        total_e = scope1 + scope2
        total_s = sinks["total_tco2e"]
        neto    = total_e + total_s

        return {
            "portada":      portada,
            "emissions":    emissions,
            "scope1":       scope1,
            "scope2":       scope2,
            "total_e":      total_e,
            "total_s":      total_s,
            "neto":         neto,
            "monthly_elec": monthly_elec,
            "sinks":        sinks,
            "ganaderia":    gan,
            "suelos":       suelos,
            "maq_det":      maq_det,
            "energia":      energia,
        }, None
    except Exception as e:
        return None, str(e)


# ─── ESTILOS CSS ──────────────────────────────────────────────────────────────
# Solo se estilizan las clases propias del dashboard (kpi-card, badges, etc.).
# El resto de los componentes nativos de Streamlit se dejan en manos del tema
# elegido por el usuario (oscuro / claro / según sistema) para que ambos modos
# se vean consistentes.
st.markdown(f"""
<style>
  /* Cards KPI */
  .kpi-card {{
    background: linear-gradient(135deg, {T['card_grad'][0]}, {T['card_grad'][1]});
    border-radius: 10px;
    padding: 18px 16px 14px;
    text-align: center;
    border: 1px solid {T['border']};
    transition: transform 0.15s;
  }}
  .kpi-card:hover {{ transform: translateY(-2px); }}
  .kpi-label {{ color: {T['text_muted']}; font-size: 11px; font-weight: 600;
               text-transform: uppercase; letter-spacing: 1.2px; margin-bottom: 6px; }}
  .kpi-value {{ font-size: 30px; font-weight: 800; line-height: 1.1; }}
  .kpi-unit  {{ color: {T['text_dim']}; font-size: 11px; margin-top: 4px; }}
  .kpi-icon  {{ font-size: 20px; margin-bottom: 4px; }}

  /* Badges de scope */
  .badge {{ border-radius: 12px; padding: 2px 10px; font-size: 11px; font-weight: 700; }}
  .badge-s1  {{ background: #C0000033; color: #FF8888; border: 1px solid #C00000; }}
  .badge-s2  {{ background: #0070C033; color: #88BBFF; border: 1px solid #0070C0; }}
  .badge-sk  {{ background: #00B05033; color: #88EE88; border: 1px solid #00B050; }}

  /* Separador de sección */
  .section-title {{ color: #81C784 !important; font-size: 14px !important; font-weight: 700 !important;
                   text-transform: uppercase; letter-spacing: 2px; margin: 20px 0 8px 0; }}
  h1 {{ color: #00B050 !important; }}
  h3 {{ color: {T['h3']} !important; }}
</style>
""", unsafe_allow_html=True)


# ─── SIDEBAR ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⚙️ Configuración")

    excel_path = st.text_input("📂 Ruta del archivo Excel", value=EXCEL_DEFAULT,
                                help="Ruta completa al archivo Matriz_HuellaCarbono_Agropecuario.xlsx")

    auto_on = st.toggle("🔄 Auto-actualizar al guardar Excel", value=True)

    if st.button("⚡ Actualizar ahora", type="primary", width="stretch"):
        st.cache_data.clear()
        st.rerun()

    st.divider()

    if os.path.exists(excel_path):
        mtime   = os.path.getmtime(excel_path)
        last_m  = datetime.fromtimestamp(mtime).strftime("%d/%m/%Y  %H:%M:%S")
        size_kb = os.path.getsize(excel_path) / 1024
        st.success("Archivo encontrado ✓")
        st.markdown(f"**Guardado:** `{last_m}`")
        st.markdown(f"**Tamaño:** `{size_kb:.0f} KB`")
    else:
        st.error("Archivo no encontrado")
        mtime = 0

    st.divider()
    st.markdown("""
**Cómo funciona:**
1. Completar celdas 🟡 en la Matriz Excel
2. Guardar el archivo `Ctrl+S`
3. El dashboard detecta el cambio y se actualiza solo

---
🟡 **Amarillo** → ingresar datos
🟢 **Verde** → calculado auto
🔵 **Celeste** → factor de referencia

---
📋 GHG Protocol + ISO 14064-1
🧮 IPCC 2006 | AR6 (CH4 fósil 29,8 / biogénico 27,2)
⚡ Factor CAMMESA 0,383 (2023) — verificar valor vigente cada año
""")


# ─── VERIFICACIÓN DE ARCHIVO ──────────────────────────────────────────────────
if not os.path.exists(excel_path):
    st.error(f"No se encontró el archivo: `{excel_path}`")
    st.info("Verificá la ruta en la barra lateral. El archivo debe estar en la carpeta Downloads.")
    st.stop()


# ─── AUTO-REFRESH POR CAMBIO DE ARCHIVO ──────────────────────────────────────
if auto_on:
    current_mtime = os.path.getmtime(excel_path)
    prev_mtime    = st.session_state.get("prev_mtime", -1)
    if current_mtime != prev_mtime:
        st.session_state.prev_mtime = current_mtime
        st.cache_data.clear()


# ─── CARGA DE DATOS ───────────────────────────────────────────────────────────
mtime    = os.path.getmtime(excel_path)
data, err = load_all(mtime, excel_path)

if data is None:
    st.error(f"Error al leer el archivo Excel: {err}")
    st.warning("Asegurate de que el archivo no esté abierto en Excel al mismo tiempo (o cerrá y reabrí).")
    st.stop()

p = data["portada"]
em = data["emissions"]
total_e = data["total_e"]
neto    = data["neto"]


# ─── ENCABEZADO ───────────────────────────────────────────────────────────────
col_h1, col_h2 = st.columns([7, 3])
with col_h1:
    st.markdown("# 🌿 Dashboard — Huella de Carbono Corporativa")
    anio_str = f" | Año {p['anio']}" if p['anio'] != "—" else ""
    nombre_str = f"**{p['nombre']}**" if p['nombre'] != "—" else "_Establecimiento sin nombre_"
    st.markdown(f"{nombre_str}  ·  {p['provincia']}{anio_str}  ·  {p['ha_prod']} ha productivas")

with col_h2:
    last_upd = datetime.fromtimestamp(mtime).strftime("%d/%m/%Y %H:%M")
    st.markdown(f"""
    <div style="background:{T['bg_card']};border-radius:8px;padding:12px 16px;text-align:right;margin-top:8px;">
        <span style="color:{T['text_dim']};font-size:11px;">Última lectura del archivo</span><br>
        <span style="color:{T['h3']};font-weight:700;">{last_upd}</span>
    </div>
    """, unsafe_allow_html=True)

st.divider()


# ─── FILA 1 — KPI CARDS ──────────────────────────────────────────────────────
st.markdown('<p class="section-title">Resumen del Inventario</p>', unsafe_allow_html=True)

def kpi_card(icon, label, value, unit, border_color):
    return f"""
    <div class="kpi-card" style="border-top: 3px solid {border_color};">
        <div class="kpi-icon">{icon}</div>
        <div class="kpi-label">{label}</div>
        <div class="kpi-value" style="color:{border_color};">{value}</div>
        <div class="kpi-unit">{unit}</div>
    </div>"""

c1, c2, c3, c4, c5 = st.columns(5)
with c1:
    st.markdown(kpi_card("🏭", "Emisiones Totales", fmt_t(total_e), "tCO₂e / año", "#C00000"), unsafe_allow_html=True)
with c2:
    st.markdown(kpi_card("🔥", "Scope 1 — Directas", fmt_t(data['scope1']), "tCO₂e / año", "#E07070"), unsafe_allow_html=True)
with c3:
    st.markdown(kpi_card("⚡", "Scope 2 — Electricidad", fmt_t(data['scope2']), "tCO₂e / año", "#0070C0"), unsafe_allow_html=True)
with c4:
    st.markdown(kpi_card("🌳", "Capturas Sumideros", fmt_t(data['total_s']), "tCO₂e / año", "#00B050"), unsafe_allow_html=True)
with c5:
    net_col   = "#00B050" if neto < 0 else ("#ED7D31" if total_e > 0 else T['text_muted'])
    net_label = "tCO₂e neto / año"
    st.markdown(kpi_card("⚖️", "Huella Neta", fmt_t(neto), net_label, net_col), unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)


# ─── FILA 2 — TORTA + GAUGE ───────────────────────────────────────────────────
st.markdown('<p class="section-title">Distribución de Emisiones</p>', unsafe_allow_html=True)
col_pie, col_gauge = st.columns([3, 2])

with col_pie:
    non_zero = {k: v for k, v in em.items() if v > 0.001}
    if non_zero:
        labels = [CAT_LABELS[k] for k in non_zero]
        values = list(non_zero.values())
        fig_pie = go.Figure(go.Pie(
            labels=labels, values=values, hole=0.58,
            marker=dict(colors=COLORS_PIE[:len(labels)]),
            textinfo="label+percent",
            textfont=dict(size=11, color=T['text_soft']),
            hovertemplate="<b>%{label}</b><br>%{value:.2f} tCO₂e<br>%{percent}<extra></extra>",
            direction="clockwise",
        ))
        fig_pie.add_annotation(
            x=0.5, y=0.5, showarrow=False,
            text=f"<b>{total_e:.1f}</b><br><span style='font-size:11px'>tCO₂e</span>",
            font=dict(size=17, color=T['text']),
        )
        fig_pie.update_layout(
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            font=dict(color=T['text_soft']),
            legend=dict(orientation="v", x=1.02, y=0.5, font=dict(size=11)),
            margin=dict(l=10, r=10, t=20, b=10),
            height=330,
        )
        st.plotly_chart(fig_pie, width="stretch")
    else:
        st.info("Completar datos en la Matriz Excel para ver la distribución de emisiones.")

with col_gauge:
    pct_neutralized = 0.0
    if total_e > 0 and data["total_s"] < 0:
        pct_neutralized = min(abs(data["total_s"]) / total_e * 100, 100)

    fig_g = go.Figure(go.Indicator(
        mode="gauge+number",
        value=pct_neutralized,
        number={"suffix": "%", "font": {"size": 38, "color": COLOR_SK}},
        title={"text": "Emisiones neutralizadas<br>por sumideros propios",
               "font": {"color": T['text_muted'], "size": 12}},
        gauge={
            "axis": {"range": [0, 100], "tickfont": {"color": T['text_muted']}, "tickcolor": T['text_muted']},
            "bar":  {"color": COLOR_SK, "thickness": 0.3},
            "bgcolor": T['bg_alt'],
            "bordercolor": T['grid'],
            "steps": [
                {"range": [0,  25], "color": T['gauge_steps'][0]},
                {"range": [25, 50], "color": T['gauge_steps'][1]},
                {"range": [50, 75], "color": T['gauge_steps'][2]},
                {"range": [75,100], "color": T['gauge_steps'][3]},
            ],
            "threshold": {"line": {"color": "#FFD700", "width": 3},
                          "thickness": 0.8, "value": 100},
        }
    ))
    fig_g.update_layout(
        paper_bgcolor="rgba(0,0,0,0)",
        font=dict(color=T['text_soft']),
        height=260,
        margin=dict(l=20, r=20, t=50, b=10),
    )
    st.plotly_chart(fig_g, width="stretch")

    sk = data["sinks"]
    st.markdown(f"""
    <div style="background:{T['bg_card']};border-radius:8px;padding:12px 16px;font-size:13px;">
        🌲 Monte nativo: <b style="color:#81C784;">{fmt_t(sk['monte_tco2e'])} tCO₂e</b><br>
        🌾 Pasturas: <b style="color:#81C784;">{fmt_t(sk['pasturas_tco2e'])} tCO₂e</b><br>
        <hr style="border-color:{T['grid']};margin:8px 0;">
        ⚖️ Huella neta: <b style="color:{net_col};">{fmt_t(neto)} tCO₂e</b>
    </div>
    """, unsafe_allow_html=True)


# ─── FILA 3 — BARRAS S1 + KPIs ───────────────────────────────────────────────
st.markdown("<br>", unsafe_allow_html=True)
col_bars, col_kpis = st.columns([3, 2])

with col_bars:
    st.markdown('<p class="section-title">Detalle Scope 1 — por Fuente</p>', unsafe_allow_html=True)
    s1_items = {k: v for k in SCOPE1_KEYS if (v := em[k]) > 0.001}
    if s1_items:
        df_s1 = pd.DataFrame({
            "Fuente": [CAT_LABELS[k] for k in s1_items],
            "tCO2e":  list(s1_items.values()),
        }).sort_values("tCO2e", ascending=True)

        fig_bar = go.Figure(go.Bar(
            y=df_s1["Fuente"], x=df_s1["tCO2e"],
            orientation="h",
            marker=dict(
                color=df_s1["tCO2e"],
                colorscale=[[0, "#FF9999"], [1, "#C00000"]],
                showscale=False,
            ),
            text=[f" {v:.2f} t" for v in df_s1["tCO2e"]],
            textposition="outside",
            textfont=dict(color=T['text_soft'], size=11),
            hovertemplate="<b>%{y}</b><br>%{x:.3f} tCO₂e<extra></extra>",
        ))
        fig_bar.update_layout(
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            xaxis=dict(gridcolor=T['grid'], color=T['text_muted'], title="tCO₂e / año",
                       zeroline=True, zerolinecolor=T['grid']),
            yaxis=dict(color=T['text_soft']),
            font=dict(color=T['text_soft']),
            margin=dict(l=10, r=70, t=10, b=40),
            height=300,
        )
        st.plotly_chart(fig_bar, width="stretch")
    else:
        st.info("Sin datos de Scope 1 todavía.")

with col_kpis:
    st.markdown('<p class="section-title">KPIs de Intensidad</p>', unsafe_allow_html=True)
    ha_prod = flt(p["ha_prod"]) or 900
    kpis_data = [
        ("tCO₂e / ha productiva",   f"{neto/ha_prod:.3f}"  if total_e > 0 else "—", "#4FC3F7"),
        ("% Scope 1 / total",        f"{pct_of(data['scope1'], total_e):.1f}%"  if total_e > 0 else "—", "#E07070"),
        ("% Scope 2 / total",        f"{pct_of(data['scope2'], total_e):.1f}%"  if total_e > 0 else "—", COLOR_S2),
        ("% emisiones neutralizadas",f"{pct_neutralized:.1f}%", COLOR_SK),
        ("Capturas totales (abs)",   f"{abs(data['total_s']):.1f} t"  if data['total_s'] != 0 else "—", "#81C784"),
        ("Balance neto",             f"{fmt_t(neto)} tCO₂e", net_col),
    ]
    for label, val, color in kpis_data:
        st.markdown(f"""
        <div style="background:{T['bg_card']};border-left:3px solid {color};border-radius:0 6px 6px 0;
                    padding:9px 14px;margin:5px 0;">
            <span style="color:{T['text_muted']};font-size:11px;">{label}</span><br>
            <span style="color:{color};font-size:20px;font-weight:700;">{val}</span>
        </div>
        """, unsafe_allow_html=True)


# ─── FILA 4 — ELECTRICIDAD MENSUAL ───────────────────────────────────────────
st.markdown("<br>", unsafe_allow_html=True)
st.markdown('<p class="section-title">Scope 2 — Electricidad de Red por Mes</p>', unsafe_allow_html=True)

df_elec = pd.DataFrame(data["monthly_elec"])
if df_elec["kwh"].sum() > 0:
    fig_elec = make_subplots(specs=[[{"secondary_y": True}]])
    fig_elec.add_trace(go.Bar(
        x=df_elec["mes"], y=df_elec["kwh"],
        name="kWh consumidos", marker_color="#0070C0",
        hovertemplate="%{x}<br>kWh: %{y:,.0f}<extra></extra>",
    ), secondary_y=False)
    fig_elec.add_trace(go.Scatter(
        x=df_elec["mes"], y=df_elec["tco2e"],
        name="tCO₂e", mode="lines+markers",
        line=dict(color="#FF9900", width=2),
        marker=dict(size=7, color="#FF9900"),
        hovertemplate="%{x}<br>tCO₂e: %{y:.3f}<extra></extra>",
    ), secondary_y=True)
    fig_elec.update_xaxes(color=T['text_muted'], gridcolor=T['grid'])
    fig_elec.update_yaxes(title_text="kWh",   secondary_y=False, color=T['h3'], gridcolor=T['grid'])
    fig_elec.update_yaxes(title_text="tCO₂e", secondary_y=True,  color="#FF9900", gridcolor="rgba(0,0,0,0)")
    fig_elec.update_layout(
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color=T['text_soft']),
        legend=dict(orientation="h", y=-0.25, font=dict(size=11)),
        margin=dict(l=60, r=60, t=10, b=60),
        height=280,
        hovermode="x unified",
    )
    st.plotly_chart(fig_elec, width="stretch")

    total_kwh  = df_elec["kwh"].sum()
    total_elec = df_elec["tco2e"].sum()
    prom_kwh   = df_elec[df_elec["kwh"] > 0]["kwh"].mean()
    c1e, c2e, c3e, c4e = st.columns(4)
    with c1e: st.metric("Total kWh / año",     f"{total_kwh:,.0f} kWh")
    with c2e: st.metric("Total tCO₂e Sc.2",    f"{total_elec:.2f} tCO₂e")
    with c3e: st.metric("Promedio mensual",     f"{prom_kwh:,.0f} kWh" if prom_kwh else "—")
    with c4e: st.metric("FE utilizado",         "0,383 kg CO₂e/kWh (CAMMESA 2023)")
else:
    st.info("Cargar las facturas de electricidad en la hoja **S2-1 ELECTRICIDAD** para ver el gráfico mensual.")


# ─── FILA 5 — GANADO (si hay datos) ──────────────────────────────────────────
gan = data["ganaderia"]
if gan["ferm_detail"]:
    st.markdown("<br>", unsafe_allow_html=True)
    col_gan, col_suelos = st.columns(2)

    with col_gan:
        st.markdown('<p class="section-title">Ganadería — Fermentación Entérica por Categoría</p>',
                    unsafe_allow_html=True)
        df_gan = pd.DataFrame(gan["ferm_detail"]).sort_values("tco2e", ascending=False)
        fig_gan = go.Figure(go.Bar(
            x=df_gan["categoria"], y=df_gan["tco2e"],
            marker_color="#E05050",
            text=[f"{v:.2f}" for v in df_gan["tco2e"]],
            textposition="outside",
            textfont=dict(color=T['text_soft'], size=10),
            hovertemplate="<b>%{x}</b><br>%{y:.3f} tCO₂e<extra></extra>",
        ))
        fig_gan.update_layout(
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            xaxis=dict(color=T['text_soft'], tickangle=-25),
            yaxis=dict(gridcolor=T['grid'], color=T['text_muted'], title="tCO₂e / año"),
            font=dict(color=T['text_soft']),
            margin=dict(l=40, r=10, t=10, b=80),
            height=280,
        )
        st.plotly_chart(fig_gan, width="stretch")

    with col_suelos:
        st.markdown('<p class="section-title">Suelos Agrícolas — N₂O + CO₂ Urea por Cultivo</p>',
                    unsafe_allow_html=True)
        s_det = data["suelos"]["detail"]
        if s_det:
            df_suel = pd.DataFrame(s_det).sort_values("tco2e", ascending=False)
            fig_suel = go.Figure(go.Bar(
                x=df_suel["cultivo"], y=df_suel["tco2e"],
                marker_color="#FF8000",
                text=[f"{v:.2f}" for v in df_suel["tco2e"]],
                textposition="outside",
                textfont=dict(color=T['text_soft'], size=10),
                hovertemplate="<b>%{x}</b><br>%{y:.3f} tCO₂e<extra></extra>",
            ))
            fig_suel.update_layout(
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                xaxis=dict(color=T['text_soft'], tickangle=-25),
                yaxis=dict(gridcolor=T['grid'], color=T['text_muted'], title="tCO₂e / año"),
                font=dict(color=T['text_soft']),
                margin=dict(l=40, r=10, t=10, b=80),
                height=280,
            )
            st.plotly_chart(fig_suel, width="stretch")
        else:
            st.info("Completar cultivos en la hoja S1-4 para ver este gráfico.")


# ─── FILA 6 — TABLA RESUMEN COMPLETA ─────────────────────────────────────────
st.markdown("<br>", unsafe_allow_html=True)
st.markdown('<p class="section-title">Tabla Resumen del Inventario</p>', unsafe_allow_html=True)

scope_map = {
    "maquinaria":   "Scope 1", "energia_est": "Scope 1",
    "fermentacion": "Scope 1", "estiercol":   "Scope 1",
    "suelos":       "Scope 1", "efluentes":   "Scope 1",
    "electricidad": "Scope 2",
}
rows_table = []
for k, v in em.items():
    pct = pct_of(v, total_e)
    rows_table.append({
        "Código":       f"S{'2' if k=='electricidad' else '1'}-{'1' if k=='maquinaria' else '2' if k=='energia_est' else '3' if k=='fermentacion' else '4' if k=='estiercol' else '5' if k=='suelos' else '6' if k=='efluentes' else '1'}",
        "Categoría":    CAT_LABELS[k],
        "Scope":        scope_map[k],
        "tCO₂e / año":  round(v, 3),
        "% del total":  f"{pct:.1f}%",
    })

rows_table.append({
    "Código": "—", "Categoría": "▶ SUBTOTAL EMISIONES", "Scope": "S1+S2",
    "tCO₂e / año": round(total_e, 3), "% del total": "100,0%",
})
rows_table.append({
    "Código": "SK-1", "Categoría": "Monte Nativo", "Scope": "Sumidero",
    "tCO₂e / año": round(data["sinks"]["monte_tco2e"], 3),
    "% del total": f"{pct_of(abs(data['sinks']['monte_tco2e']), total_e):.1f}%",
})
rows_table.append({
    "Código": "SK-2", "Categoría": "Pasturas", "Scope": "Sumidero",
    "tCO₂e / año": round(data["sinks"]["pasturas_tco2e"], 3),
    "% del total": f"{pct_of(abs(data['sinks']['pasturas_tco2e']), total_e):.1f}%",
})
rows_table.append({
    "Código": "—", "Categoría": "▶▶ HUELLA NETA", "Scope": "Neto",
    "tCO₂e / año": round(neto, 3), "% del total": "—",
})

df_table = pd.DataFrame(rows_table)
st.dataframe(
    df_table,
    width="stretch",
    hide_index=True,
    column_config={
        "tCO₂e / año": st.column_config.NumberColumn("tCO₂e / año", format="%.3f"),
    },
)


# ─── FOOTER ───────────────────────────────────────────────────────────────────
st.divider()
fc1, fc2, fc3, fc4 = st.columns(4)
with fc1: st.caption("📋 GHG Protocol Corporate Standard")
with fc2: st.caption("📏 ISO 14064-1:2018")
with fc3: st.caption("🧮 IPCC 2006 | CAMMESA 2023 (verificar) | AR6 (CH4 fósil 29,8 / biogénico 27,2)")
with fc4: st.caption(f"🕐 Última renderización: {datetime.now().strftime('%H:%M:%S')}")


# ─── AUTO-REFRESH PERIÓDICO (cada 10 s si está activado) ─────────────────────
if auto_on:
    time.sleep(10)
    if os.path.exists(excel_path):
        new_mtime = os.path.getmtime(excel_path)
        if new_mtime != st.session_state.get("prev_mtime", -1):
            st.cache_data.clear()
    st.rerun()
