#!/usr/bin/env python3
"""
Dashboard interactivo — Huella de Carbono Vitivinícola
Lee la planilla Matriz_HuellaCarbono_Vitivinicola.xlsx y recalcula en vivo.

Ejecutar:  streamlit run dashboard_vino.py
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import openpyxl
import os
import time
from datetime import datetime

st.set_page_config(
    page_title="Huella de Carbono — Bodega",
    page_icon="🍷",
    layout="wide",
)

DARK = (st.context.theme.type or "dark") != "light"
FG = "#FAFAFA" if DARK else "#31333F"
BG_CARD = "#262730" if DARK else "#F0F2F6"

C_S1, C_S2, C_S3, C_BIO = "#C00000", "#0070C0", "#7030A0", "#548235"

# Orden de búsqueda: copia "viva" en Downloads (trabajo local con un cliente),
# luego el EJEMPLO del repo (Streamlit Cloud / demo), luego la planilla en blanco.
_REPO_DIR = os.path.dirname(os.path.abspath(__file__))
_CANDIDATOS = [
    os.path.expanduser("~/Downloads/Matriz_HuellaCarbono_Vitivinicola.xlsx"),
    os.path.join(_REPO_DIR, "Matriz_HuellaCarbono_Vitivinicola_EJEMPLO.xlsx"),
    os.path.join(_REPO_DIR, "Matriz_HuellaCarbono_Vitivinicola.xlsx"),
]
EXCEL_DEFAULT = next((p for p in _CANDIDATOS if os.path.exists(p)), _CANDIDATOS[1])

# ─── Factores (espejo de la planilla — FACTORES EMISION) ────────────────────
FE_GASOIL, FE_NAFTA, FE_GN, FE_GLP = 2.68, 2.31, 2.04, 3.02
FE_N, FE_UREA, FE_CAL, FE_DOLOMITA = 4.29, 0.733, 0.440, 0.477
GWP = {"R-134a": 1530, "R-404A": 4728, "R-410A": 2256, "R-22": 1960, "Amoníaco (NH3)": 0}
FE_ELEC = 0.383
FE_VIDRIO = 1.2
FE_CIERRE = {"Corcho natural": 0.00183, "Tapón sintético": 0.0148, "Tapa rosca (aluminio)": 0.0372}
FE_ALU, FE_PAPEL, FE_CARTON = 8.6, 1.1, 0.94
FE_FERT_FAB, FE_AGROQ_FAB = 1.2, 6.3
FE_CAMION, FE_TREN, FE_BARCO = 0.062, 0.022, 0.012
FE_FERMENT = 0.098  # kg CO2 biogénico / L vino


def flt(val):
    try:
        if val is None or val == "":
            return 0.0
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def fmt_t(val, dec=1):
    return f"{val:,.{dec}f}".replace(",", "X").replace(".", ",").replace("X", ".")


# ─── Lectores por hoja (leen celdas de INGRESO y recalculan) ─────────────────

def read_maquinaria(wb):
    """S1 MAQUINARIA — filas 3-8: D=consumo, F=FE"""
    ws = wb["S1 MAQUINARIA"]
    rows, total = [], 0.0
    for r in range(3, 9):
        nombre = str(ws.cell(r, 1).value or "")
        dato = flt(ws.cell(r, 4).value)
        fe = flt(ws.cell(r, 6).value)
        t = dato * fe / 1000
        if dato:
            rows.append({"item": nombre, "tco2e": t})
        total += t
    return rows, total


def read_energia_fija(wb):
    """S1 ENERGIA FIJA — filas 3-6"""
    ws = wb["S1 ENERGIA FIJA"]
    rows, total = [], 0.0
    for r in range(3, 7):
        nombre = str(ws.cell(r, 1).value or "")
        dato = flt(ws.cell(r, 4).value)
        fe = flt(ws.cell(r, 6).value)
        t = dato * fe / 1000
        if dato:
            rows.append({"item": nombre, "tco2e": t})
        total += t
    return rows, total


def read_vinedo(wb):
    """S1 VIÑEDO — filas 3-6: C=dato, E=FE"""
    ws = wb["S1 VIÑEDO"]
    rows, total = [], 0.0
    for r in range(3, 7):
        nombre = str(ws.cell(r, 1).value or "")
        dato = flt(ws.cell(r, 3).value)
        fe = flt(ws.cell(r, 5).value)
        t = dato * fe / 1000
        if dato:
            rows.append({"item": nombre, "tco2e": t})
        total += t
    return rows, total


def read_refrigerantes(wb):
    """S1 REFRIGERANTES — filas 3-7: C=carga, D=%fuga, E=GWP"""
    ws = wb["S1 REFRIGERANTES"]
    rows, total = [], 0.0
    for r in range(3, 8):
        gas = str(ws.cell(r, 1).value or "")
        carga = flt(ws.cell(r, 3).value)
        fuga = flt(ws.cell(r, 4).value)
        gwp = flt(ws.cell(r, 5).value)
        t = carga * fuga / 100 * gwp / 1000
        if carga:
            rows.append({"item": gas, "tco2e": t})
        total += t
    return rows, total


def read_electricidad(wb):
    """S2 ELECTRICIDAD — fila 3: C=kWh, E=FE"""
    ws = wb["S2 ELECTRICIDAD"]
    kwh = flt(ws.cell(3, 3).value)
    fe = flt(ws.cell(3, 5).value)
    solar = flt(ws.cell(4, 3).value)
    return kwh, solar, kwh * fe / 1000


def read_envases(wb):
    """S3 ENVASES — botellas filas 4-6, packaging filas 10-15"""
    ws = wb["S3 ENVASES"]
    unid = flt(ws.cell(4, 3).value)
    peso_g = flt(ws.cell(5, 3).value)
    pct_rec = flt(ws.cell(6, 3).value)
    t_botellas = unid * peso_g / 1000 * FE_VIDRIO * (1 - 0.002 * pct_rec) / 1000
    rows = []
    if unid:
        rows.append({"item": "Botellas de vidrio", "tco2e": t_botellas})
    total = t_botellas
    for r in range(10, 16):
        nombre = str(ws.cell(r, 1).value or "")
        dato = flt(ws.cell(r, 3).value)
        fe = flt(ws.cell(r, 5).value)
        t = dato * fe / 1000
        if dato:
            rows.append({"item": nombre, "tco2e": t})
        total += t
    return rows, total, unid


def read_insumos(wb):
    """S3 INSUMOS — filas 3-5"""
    ws = wb["S3 INSUMOS"]
    rows, total = [], 0.0
    for r in range(3, 6):
        nombre = str(ws.cell(r, 1).value or "")
        dato = flt(ws.cell(r, 3).value)
        fe = flt(ws.cell(r, 5).value)
        t = dato * fe / 1000
        if dato and fe:
            rows.append({"item": nombre, "tco2e": t})
        total += t
    return rows, total


def read_transporte(wb):
    """S3 TRANSPORTE — filas 3-7"""
    ws = wb["S3 TRANSPORTE"]
    rows, total = [], 0.0
    for r in range(3, 8):
        nombre = str(ws.cell(r, 1).value or "")
        dato = flt(ws.cell(r, 3).value)
        fe = flt(ws.cell(r, 5).value)
        t = dato * fe / 1000
        if dato:
            rows.append({"item": nombre, "tco2e": t})
        total += t
    return rows, total


def read_fermentacion(wb):
    ws = wb["FERMENTACION (INFO)"]
    litros = flt(ws.cell(3, 3).value)
    return litros, litros * FE_FERMENT / 1000


def read_portada(wb):
    ws = wb["PORTADA"]
    return {
        "bodega": ws.cell(5, 2).value or "Bodega (sin nombre en PORTADA)",
        "litros": flt(ws.cell(10, 2).value),
        "botellas": flt(ws.cell(11, 2).value),
        "anio": ws.cell(12, 2).value or "—",
    }


@st.cache_data(ttl=8)
def load_all(_mtime, path):
    wb = openpyxl.load_workbook(path, data_only=False)
    maq_rows, s1_maq = read_maquinaria(wb)
    fija_rows, s1_fija = read_energia_fija(wb)
    vin_rows, s1_vin = read_vinedo(wb)
    ref_rows, s1_ref = read_refrigerantes(wb)
    kwh, solar, s2 = read_electricidad(wb)
    env_rows, s3_env, unid_bot = read_envases(wb)
    ins_rows, s3_ins = read_insumos(wb)
    tra_rows, s3_tra = read_transporte(wb)
    litros_ferm, t_bio = read_fermentacion(wb)
    portada = read_portada(wb)
    s1 = s1_maq + s1_fija + s1_vin + s1_ref
    s3 = s3_env + s3_ins + s3_tra
    return {
        "portada": portada,
        "s1": s1, "s2": s2, "s3": s3, "total": s1 + s2 + s3,
        "bio": t_bio, "litros_ferm": litros_ferm,
        "kwh": kwh, "solar": solar, "unid_bot": unid_bot,
        "categorias": [
            ("Combustión móvil", s1_maq, C_S1),
            ("Combustión estacionaria", s1_fija, C_S1),
            ("Viñedo — N2O y enmiendas", s1_vin, C_S1),
            ("Refrigerantes", s1_ref, C_S1),
            ("Electricidad de red", s2, C_S2),
            ("Envases y embalajes", s3_env, C_S3),
            ("Insumos (fabricación)", s3_ins, C_S3),
            ("Transporte y distribución", s3_tra, C_S3),
        ],
        "detalle_envases": env_rows,
    }


# ─── UI ──────────────────────────────────────────────────────────────────────

st.markdown(
    f"<h2 style='margin-bottom:0'>🍷 Huella de Carbono — Bodega</h2>"
    f"<p style='color:gray;margin-top:2px'>GHG Protocol + ISO 14064-1 + Protocolo OIV GEI "
    f"| Alcances 1 + 2 + 3 | GWP IPCC AR6</p>",
    unsafe_allow_html=True,
)

with st.sidebar:
    st.markdown("### ⚙️ Configuración")
    excel_path = st.text_input("📂 Ruta del archivo Excel", value=EXCEL_DEFAULT)
    if st.button("⚡ Actualizar ahora", type="primary", width="stretch"):
        st.cache_data.clear()
        st.rerun()
    st.divider()
    if os.path.exists(excel_path):
        mtime = os.path.getmtime(excel_path)
        st.success("Archivo encontrado ✓")
        st.markdown(f"**Guardado:** `{datetime.fromtimestamp(mtime):%d/%m/%Y %H:%M}`")
    else:
        st.error("Archivo no encontrado")
    st.divider()
    st.markdown(
        "🟡 Celdas amarillas = datos de la bodega\n\n"
        "🟦 Celestes = factores de referencia\n\n"
        "El dashboard recalcula al guardar el Excel."
    )

if not os.path.exists(excel_path):
    st.error(f"No se encontró el archivo: `{excel_path}`")
    st.stop()

try:
    data = load_all(os.path.getmtime(excel_path), excel_path)
except Exception as err:
    st.error(f"Error al leer el Excel: {err}")
    st.warning("Cerrá el archivo en Excel si lo tenés abierto y reintentá.")
    st.stop()

p = data["portada"]
st.markdown(f"**{p['bodega']}** — Inventario {p['anio']}")

# KPIs
kg_botella = data["total"] * 1000 / p["botellas"] if p["botellas"] else 0
k1, k2, k3, k4, k5 = st.columns(5)
k1.metric("TOTAL (tCO2e/año)", fmt_t(data["total"]))
k2.metric("Alcance 1", fmt_t(data["s1"]))
k3.metric("Alcance 2", fmt_t(data["s2"]))
k4.metric("Alcance 3", fmt_t(data["s3"]))
k5.metric("kg CO2e / botella", fmt_t(kg_botella, 2),
          help="Indicador organizacional. Benchmark bibliográfico: 0,9–1,9 kg/botella")

st.divider()
col_a, col_b = st.columns([1, 1.4])

with col_a:
    st.markdown("#### Participación por alcance")
    fig_pie = go.Figure(go.Pie(
        labels=["Alcance 1", "Alcance 2", "Alcance 3"],
        values=[data["s1"], data["s2"], data["s3"]],
        hole=0.55,
        marker=dict(colors=[C_S1, C_S2, C_S3]),
        textinfo="label+percent",
    ))
    fig_pie.update_layout(showlegend=False, height=330,
                          margin=dict(t=10, b=10, l=10, r=10),
                          paper_bgcolor="rgba(0,0,0,0)", font_color=FG)
    st.plotly_chart(fig_pie, width="stretch")
    st.info(f"🌱 **CO2 biogénico de fermentación: {fmt_t(data['bio'])} tCO2/año** — "
            "se reporta por separado y no suma al total (criterio OIV/FIVS).")

with col_b:
    st.markdown("#### Emisiones por categoría (tCO2e/año)")
    cats = [c for c in data["categorias"]]
    cats.sort(key=lambda x: x[1])
    fig_bar = go.Figure(go.Bar(
        x=[c[1] for c in cats],
        y=[c[0] for c in cats],
        orientation="h",
        marker_color=[c[2] for c in cats],
        text=[fmt_t(c[1]) for c in cats],
        textposition="outside",
    ))
    fig_bar.update_layout(height=380, margin=dict(t=10, b=10, l=10, r=40),
                          paper_bgcolor="rgba(0,0,0,0)",
                          plot_bgcolor="rgba(0,0,0,0)", font_color=FG,
                          xaxis_title="tCO2e/año")
    st.plotly_chart(fig_bar, width="stretch")

# Detalle de envases (la fuente dominante)
if data["detalle_envases"]:
    st.markdown("#### 🍾 Detalle de envases y embalajes — la fuente dominante en vino")
    df_env = pd.DataFrame(data["detalle_envases"])
    df_env = df_env.sort_values("tco2e", ascending=False)
    fig_env = go.Figure(go.Bar(
        x=[r for r in df_env["item"]],
        y=[v for v in df_env["tco2e"]],
        marker_color=C_S3,
        text=[fmt_t(v, 2) for v in df_env["tco2e"]],
        textposition="outside",
    ))
    fig_env.update_layout(height=320, margin=dict(t=10, b=10, l=10, r=10),
                          paper_bgcolor="rgba(0,0,0,0)",
                          plot_bgcolor="rgba(0,0,0,0)", font_color=FG,
                          yaxis_title="tCO2e/año")
    st.plotly_chart(fig_env, width="stretch")
    st.caption("Palancas típicas de reducción: botella más liviana, mayor % de vidrio reciclado "
               "y optimización del cierre. El vidrio suele explicar ~40% de la huella total de una bodega.")

st.divider()
st.caption("Matriz de Huella de Carbono Vitivinícola — ESG Consulting | "
           "Factores: IPCC 2006/2019, GWP AR6, CAMMESA, PwC-Amorim, Glass Alliance Europe. "
           "Indicador por botella: organizacional (no certificable ISO 14067).")
