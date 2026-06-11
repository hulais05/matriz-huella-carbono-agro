#!/usr/bin/env python3
"""
Genera la Matriz de Recopilación de Datos — Huella de Carbono
Sector Vitivinícola | GHG Protocol + ISO 14064-1 + Protocolo OIV GEI

Alcances cubiertos: Scope 1 + Scope 2 + Scope 3 (envases, insumos, transporte)
CO2 de fermentación: biogénico, se reporta por separado (no suma al total),
conforme a la práctica del protocolo OIV y la calculadora FIVS/IWCC.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── HELPERS (mismos de la matriz agropecuaria) ──────────────────────────────

def brd(style='thin'):
    s = Side(style=style, color='FF000000')
    return Border(left=s, right=s, top=s, bottom=s)

def fl(hex6):
    return PatternFill(fill_type='solid', fgColor='FF' + hex6)

def set_cell(ws, r, c, value=None, bg='F2F2F2', fg='000000',
             bold=False, sz=10, halign='left', valign='center',
             wrap=True, italic=False, border_style='thin'):
    cell = ws.cell(row=r, column=c, value=value)
    cell.fill = fl(bg)
    cell.font = Font(bold=bold, color='FF' + fg, size=sz, italic=italic)
    cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    cell.border = brd(border_style)
    return cell

def header(ws, r, c, text, bg='1F3864', fg='FFFFFF', sz=10, bold=True, halign='center'):
    return set_cell(ws, r, c, text, bg=bg, fg=fg, bold=bold, sz=sz, halign=halign)

def inp(ws, r, c, value=None):
    """Celda de ingreso — amarillo"""
    return set_cell(ws, r, c, value, bg='FFFF99', fg='000000', halign='center')

def calc(ws, r, c, formula):
    """Celda calculada — verde"""
    return set_cell(ws, r, c, formula, bg='E2EFDA', fg='375623', italic=True, halign='center')

def ref(ws, r, c, value):
    """Factor de referencia — celeste"""
    return set_cell(ws, r, c, value, bg='BDD7EE', fg='1F3864', halign='center')

def lbl(ws, r, c, text, bg='F2F2F2', bold=False):
    return set_cell(ws, r, c, text, bg=bg, fg='000000', bold=bold, halign='left')

def merge(ws, r1, c1, r2, c2, text, bg='1F3864', fg='FFFFFF',
          bold=True, sz=11, halign='center'):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=text)
    cell.fill = fl(bg)
    cell.font = Font(bold=bold, color='FF' + fg, size=sz)
    cell.alignment = Alignment(horizontal=halign, vertical='center', wrap_text=True)
    cell.border = brd()
    return cell

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def nota(ws, r, c1, c2, text):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(row=r, column=c1, value=text)
    cell.fill = fl('FFF2CC')
    cell.font = Font(size=8, italic=True, color='FF7F6000')
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell.border = brd()

# ─── COLORES ─────────────────────────────────────────────────────────────────
NAVY  = '1F3864'
TITLE = '4C0F2E'   # bordó vitivinícola
S1_D  = 'C00000'; S1_L = 'FFC7CE'   # Scope 1
S2_D  = '0070C0'; S2_L = 'DCE6F1'   # Scope 2
S3_D  = '7030A0'; S3_L = 'E4DFEC'   # Scope 3
BIO_D = '548235'; BIO_L = 'E2EFDA'  # Biogénico (informativo)
OR_D  = 'ED7D31'; OR_L  = 'FCE4D6'  # Resumen
GR1   = 'F2F2F2'

# Celdas de subtotal por hoja (para vincular el RESUMEN) — se completan al construir
SUBTOT = {}

wb = openpyxl.Workbook()

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 1: PORTADA
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "PORTADA"
ws.sheet_properties.tabColor = TITLE
set_col_widths(ws, [38, 16, 16, 16, 16, 30, 14, 14, 14])

merge(ws, 1, 1, 1, 9, "MATRIZ DE RECOPILACIÓN DE DATOS — HUELLA DE CARBONO CORPORATIVA",
      TITLE, sz=16)
merge(ws, 2, 1, 2, 9,
      "GHG Protocol (Corporate Standard) + ISO 14064-1:2018 + Protocolo OIV GEI  |  Sector Vitivinícola",
      TITLE, sz=11)
ws.row_dimensions[1].height = 42
ws.row_dimensions[2].height = 26

merge(ws, 4, 1, 4, 5, "IDENTIFICACIÓN DE LA BODEGA", NAVY, sz=11)
id_fields = [
    ("Razón Social / Nombre de la bodega", None),
    ("CUIT", None),
    ("Provincia / Departamento", None),
    ("Dirección de la bodega", None),
    ("Superficie de viñedo propio (ha)", 25),
    ("Producción anual de vino (litros)", 180000),
    ("Botellas producidas por año (0,75 L)", 240000),
    ("Año del inventario", None),
    ("Año base del inventario", None),
    ("Responsable del inventario", None),
    ("Correo electrónico", None),
    ("Teléfono de contacto", None),
]
for i, (label, val) in enumerate(id_fields):
    r = 5 + i
    lbl(ws, r, 1, label, GR1, bold=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    inp(ws, r, 2, val)
    for c in range(3, 6):
        ws.cell(row=r, column=c).fill = fl('FFFF99')
        ws.cell(row=r, column=c).border = brd()

merge(ws, 4, 6, 4, 9, "LÍMITES DEL INVENTARIO", NAVY, sz=11)
lim_fields = [
    ("Enfoque organizacional", "Control Operacional"),
    ("Alcance cubierto", "Scope 1 + 2 + 3 (envases, insumos, transporte)"),
    ("CO2 de fermentación", "Biogénico — reportado por separado (OIV/FIVS)"),
    ("GWP utilizados", "IPCC AR6 (2021) — horizonte 100 años"),
    ("FE electricidad", "Factor Red Argentina (CAMMESA)"),
    ("Factor red eléctrica (kg CO2e/kWh)", "0,383 (CAMMESA 2023 — verificar valor vigente)"),
    ("Unidad funcional de intensidad", "Botella de 0,75 L (intensidad organizacional)"),
    ("Exclusiones", "Uso y fin de vida del producto (cradle-to-gate)"),
]
for i, (label, val) in enumerate(lim_fields):
    r = 5 + i
    lbl(ws, r, 6, label, GR1, bold=True)
    ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=9)
    ref(ws, r, 7, val)
    for c in range(8, 10):
        ws.cell(row=r, column=c).fill = fl('BDD7EE')
        ws.cell(row=r, column=c).border = brd()

merge(ws, 18, 1, 18, 9, "FUENTES DE EMISIÓN DECLARADAS", NAVY, sz=11)
fuentes = [
    ("S1-1", "Combustión móvil", "Tractores de viñedo, pulverizadora, utilitarios y camión propio"),
    ("S1-2", "Combustión estacionaria", "Caldera, generador y agua caliente sanitaria (gasoil / GN / GLP)"),
    ("S1-3", "Viñedo — N2O de suelos", "Fertilización nitrogenada, urea y enmiendas calcáreas"),
    ("S1-4", "Refrigerantes", "Fugas de gases en frío de bodega (tanques, cámaras, aire acond.)"),
    ("S2-1", "Electricidad de red", "Bombas de frío, prensas, embotellado, iluminación, riego"),
    ("S3-1", "Envases y embalajes", "Botellas, tapones, cápsulas, etiquetas y cajas (fabricación)"),
    ("S3-2", "Insumos upstream", "Fertilizantes y agroquímicos (fabricación); enológicos (registro)"),
    ("S3-3", "Transporte y distribución", "Insumos a bodega, distribución nacional y exportación"),
    ("INFO", "CO2 de fermentación", "Biogénico — se cuantifica y reporta fuera del total"),
]
for j, htxt in enumerate(["Código", "Fuente", "Descripción"]):
    header(ws, 19, j + 1, htxt)
ws.merge_cells(start_row=19, start_column=3, end_row=19, end_column=9)
for i, (code, name, desc) in enumerate(fuentes):
    r = 20 + i
    bg = S1_L if code.startswith('S1') else (S2_L if code.startswith('S2')
         else (S3_L if code.startswith('S3') else BIO_L))
    set_cell(ws, r, 1, code, bg=bg, bold=True, halign='center')
    set_cell(ws, r, 2, name, bg=bg, bold=True)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=9)
    lbl(ws, r, 3, desc, GR1)
    for c in range(4, 10):
        ws.cell(row=r, column=c).fill = fl(GR1)
        ws.cell(row=r, column=c).border = brd()

nota(ws, 30, 1, 9,
     "Celdas AMARILLAS: dato a relevar en la bodega. Celdas CELESTES: factor de referencia (no modificar). "
     "Celdas VERDES: cálculo automático. Las emisiones se expresan en tCO2e/año con GWP del IPCC AR6.")

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 2: S1 MAQUINARIA (combustión móvil)
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("S1 MAQUINARIA")
ws.sheet_properties.tabColor = S1_D
set_col_widths(ws, [30, 34, 16, 14, 14, 14, 14])

merge(ws, 1, 1, 1, 7, "SCOPE 1 — COMBUSTIÓN MÓVIL  |  Maquinaria de viñedo y vehículos propios", S1_D, sz=13)
for j, htxt in enumerate(["Equipo", "Descripción / uso típico", "Combustible",
                          "Consumo anual", "Unidad", "FE (kg CO2e/un.)", "tCO2e/año"]):
    header(ws, 2, j + 1, htxt)

rows = [
    ("Tractores de viñedo",        "Labores, desbrozado, acarreo de cosecha", "Gasoil", "L/año", 2.68),
    ("Pulverizadora / atomizadora","Aplicaciones fitosanitarias",             "Gasoil", "L/año", 2.68),
    ("Cosechadora propia",         "Solo si es de la bodega (si es contratada va en S3)", "Gasoil", "L/año", 2.68),
    ("Camión propio",              "Movimiento de uva e insumos",             "Gasoil", "L/año", 2.68),
    ("Utilitarios / camionetas",   "Logística interna y comercial",           "Nafta",  "L/año", 2.31),
    ("Autoelevador / otros",       "Movimiento de pallets en bodega",         "Gasoil/GLP", "L o kg/año", 2.68),
]
for i, (eq, desc, comb, un, fe) in enumerate(rows):
    r = 3 + i
    lbl(ws, r, 1, eq, S1_L, bold=True)
    lbl(ws, r, 2, desc)
    set_cell(ws, r, 3, comb, halign='center')
    inp(ws, r, 4)
    set_cell(ws, r, 5, un, halign='center')
    ref(ws, r, 6, fe)
    calc(ws, r, 7, f'=IF(D{r}="",0,D{r}*F{r}/1000)')

r_sub = 9
merge(ws, r_sub, 1, r_sub, 6, "SUBTOTAL S1 — COMBUSTIÓN MÓVIL (tCO2e/año)", S1_D)
calc(ws, r_sub, 7, "=SUM(G3:G8)")
SUBTOT['S1 MAQUINARIA'] = f"G{r_sub}"

nota(ws, 10, 1, 7,
     "FE combinados CO2+CH4+N2O por litro — IPCC 2006 Vol. 2 (gasoil 2,68 / nafta 2,31 kg CO2e/L). "
     "Si se usa biodiesel B5/B10, ajustar FE. Fuente de datos: facturas de combustible o planilla de control.")

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 3: S1 ENERGIA FIJA (combustión estacionaria)
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("S1 ENERGIA FIJA")
ws.sheet_properties.tabColor = S1_D
set_col_widths(ws, [30, 34, 16, 14, 14, 14, 14])

merge(ws, 1, 1, 1, 7, "SCOPE 1 — COMBUSTIÓN ESTACIONARIA  |  Caldera, generador y agua caliente", S1_D, sz=13)
for j, htxt in enumerate(["Equipo", "Descripción / uso típico", "Combustible",
                          "Consumo anual", "Unidad", "FE (kg CO2e/un.)", "tCO2e/año"]):
    header(ws, 2, j + 1, htxt)

rows = [
    ("Generador eléctrico",  "Respaldo o autogeneración",          "Gasoil",      "L/año",  2.68),
    ("Caldera / calefacción","Agua caliente, calefacción de naves","Gas natural", "m3/año", 2.04),
    ("GLP (zeppelin/garrafas)","Cocina, agua caliente, autoelevador","GLP",       "kg/año", 3.02),
    ("Quemadores / otros",   "Otros equipos fijos a combustible",  "Gasoil",      "L/año",  2.68),
]
for i, (eq, desc, comb, un, fe) in enumerate(rows):
    r = 3 + i
    lbl(ws, r, 1, eq, S1_L, bold=True)
    lbl(ws, r, 2, desc)
    set_cell(ws, r, 3, comb, halign='center')
    inp(ws, r, 4)
    set_cell(ws, r, 5, un, halign='center')
    ref(ws, r, 6, fe)
    calc(ws, r, 7, f'=IF(D{r}="",0,D{r}*F{r}/1000)')

r_sub = 7
merge(ws, r_sub, 1, r_sub, 6, "SUBTOTAL S1 — COMBUSTIÓN ESTACIONARIA (tCO2e/año)", S1_D)
calc(ws, r_sub, 7, "=SUM(G3:G6)")
SUBTOT['S1 ENERGIA FIJA'] = f"G{r_sub}"

nota(ws, 8, 1, 7,
     "FE: IPCC 2006 Vol. 2 — gas natural 2,04 kg CO2e/m3, GLP 3,02 kg CO2e/kg, gasoil 2,68 kg CO2e/L.")

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 4: S1 VIÑEDO (N2O de suelos + enmiendas)
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("S1 VIÑEDO")
ws.sheet_properties.tabColor = S1_D
set_col_widths(ws, [30, 34, 16, 14, 16, 14, 14])

merge(ws, 1, 1, 1, 7, "SCOPE 1 — VIÑEDO  |  N2O de fertilización nitrogenada y enmiendas", S1_D, sz=13)
for j, htxt in enumerate(["Aplicación", "Descripción", "Dato anual", "Unidad",
                          "FE (kg CO2e/un.)", "Gas", "tCO2e/año"]):
    header(ws, 2, j + 1, htxt)

rows = [
    ("Nitrógeno aplicado (todo fert. N)", "kg de N elemental aplicados al viñedo", "kg N/año", 4.29, "N2O"),
    ("Urea — CO2 de hidrólisis",          "Solo si se aplica urea (además del N de arriba)", "kg urea/año", 0.733, "CO2"),
    ("Cal agrícola (encalado)",           "Corrección de suelos", "kg/año", 0.440, "CO2"),
    ("Dolomita",                          "Corrección de suelos", "kg/año", 0.477, "CO2"),
]
for i, (ap, desc, un, fe, gas) in enumerate(rows):
    r = 3 + i
    lbl(ws, r, 1, ap, S1_L, bold=True)
    lbl(ws, r, 2, desc)
    inp(ws, r, 3)
    set_cell(ws, r, 4, un, halign='center')
    ref(ws, r, 5, fe)
    set_cell(ws, r, 6, gas, halign='center')
    calc(ws, r, 7, f'=IF(C{r}="",0,C{r}*E{r}/1000)')

r_sub = 7
merge(ws, r_sub, 1, r_sub, 6, "SUBTOTAL S1 — VIÑEDO (tCO2e/año)", S1_D)
calc(ws, r_sub, 7, "=SUM(G3:G6)")
SUBTOT['S1 VIÑEDO'] = f"G{r_sub}"

nota(ws, 8, 1, 7,
     "N2O directo Tier 1: EF1 = 0,01 kg N2O-N/kg N (IPCC 2019 Refinement) x 44/28 x GWP 273 (AR6) "
     "= 4,29 kg CO2e/kg N. Urea: 0,20 kg C/kg x 44/12 = 0,733 kg CO2/kg (IPCC 2006). "
     "Cal 0,44 / dolomita 0,477 kg CO2/kg (IPCC 2006 Vol. 4). No incluye N2O indirecto (volatilización/lixiviación).")

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 5: S1 REFRIGERANTES (fugas de gases de frío)
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("S1 REFRIGERANTES")
ws.sheet_properties.tabColor = S1_D
set_col_widths(ws, [16, 38, 16, 16, 14, 14])

merge(ws, 1, 1, 1, 6, "SCOPE 1 — REFRIGERANTES  |  Fugas en frío de bodega (control de fermentación, cámaras)", S1_D, sz=13)
for j, htxt in enumerate(["Gas", "Equipo típico en bodega", "Carga total (kg)",
                          "Fuga anual (%)", "GWP (AR6)", "tCO2e/año"]):
    header(ws, 2, j + 1, htxt)

rows = [
    ("R-134a", "Chillers de control de temperatura de fermentación", 1530),
    ("R-404A", "Cámaras de frío / conservación",                     4728),
    ("R-410A", "Aire acondicionado de naves y oficinas",             2256),
    ("R-22",   "Equipos antiguos (en retiro — Protocolo de Montreal)", 1960),
    ("Amoníaco (NH3)", "Sistemas industriales grandes — GWP nulo",   0),
]
for i, (gas, eq, gwp) in enumerate(rows):
    r = 3 + i
    lbl(ws, r, 1, gas, S1_L, bold=True)
    lbl(ws, r, 2, eq)
    inp(ws, r, 3)
    inp(ws, r, 4, 10)
    ref(ws, r, 5, gwp)
    calc(ws, r, 6, f'=IF(C{r}="",0,C{r}*D{r}/100*E{r}/1000)')

r_sub = 8
merge(ws, r_sub, 1, r_sub, 5, "SUBTOTAL S1 — REFRIGERANTES (tCO2e/año)", S1_D)
calc(ws, r_sub, 6, "=SUM(F3:F7)")
SUBTOT['S1 REFRIGERANTES'] = f"F{r_sub}"

nota(ws, 9, 1, 6,
     "Método de tasa de fuga: carga del equipo (kg) x % de fuga anual x GWP. Tasa por defecto 10% "
     "(rango típico 5-15% en equipos comerciales — GHG Protocol, herramienta HFC). GWP AR6 a 100 años. "
     "La carga figura en la placa del equipo o en la factura de recarga del técnico de frío.")

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 6: FERMENTACION (CO2 biogénico — informativo, fuera del total)
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("FERMENTACION (INFO)")
ws.sheet_properties.tabColor = BIO_D
set_col_widths(ws, [34, 30, 16, 14, 16, 14])

merge(ws, 1, 1, 1, 6, "CO2 BIOGÉNICO DE FERMENTACIÓN ALCOHÓLICA — SE REPORTA POR SEPARADO", BIO_D, sz=13)
for j, htxt in enumerate(["Concepto", "Descripción", "Dato anual", "Unidad",
                          "Factor (kg CO2/L)", "tCO2 biogénico/año"]):
    header(ws, 2, j + 1, htxt)

r = 3
lbl(ws, r, 1, "Vino elaborado en el ejercicio", BIO_L, bold=True)
lbl(ws, r, 2, "Litros que completaron fermentación alcohólica")
inp(ws, r, 3)
set_cell(ws, r, 4, "L/año", halign='center')
ref(ws, r, 5, 0.098)
calc(ws, r, 6, '=IF(C3="",0,C3*E3/1000)')
SUBTOT['FERMENTACION (INFO)'] = "F3"

nota(ws, 5, 1, 6,
     "Estequiometría: C6H12O6 -> 2 C2H5OH + 2 CO2. Para vino de ~13% vol: ~103 g etanol/L -> ~98 g CO2/L "
     "(0,098 kg CO2/L). Este CO2 es BIOGÉNICO: fue capturado por la vid en el mismo ciclo anual, por lo que "
     "el protocolo OIV y la calculadora internacional FIVS/IWCC lo excluyen del total y lo informan por separado. "
     "NO se suma en la hoja RESUMEN.")

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 7: S2 ELECTRICIDAD
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("S2 ELECTRICIDAD")
ws.sheet_properties.tabColor = S2_D
set_col_widths(ws, [34, 34, 16, 14, 16, 14])

merge(ws, 1, 1, 1, 6, "SCOPE 2 — ELECTRICIDAD DE RED  |  Enfoque basado en localización", S2_D, sz=13)
for j, htxt in enumerate(["Concepto", "Descripción", "Dato anual", "Unidad",
                          "FE (kg CO2e/kWh)", "tCO2e/año"]):
    header(ws, 2, j + 1, htxt)

r = 3
lbl(ws, r, 1, "Electricidad comprada a la red", S2_L, bold=True)
lbl(ws, r, 2, "Frío de bodega, prensas, bombas, embotellado, riego, iluminación")
inp(ws, r, 3)
set_cell(ws, r, 4, "kWh/año", halign='center')
ref(ws, r, 5, 0.383)
calc(ws, r, 6, '=IF(C3="",0,C3*E3/1000)')

r = 4
lbl(ws, r, 1, "Solar fotovoltaica autogenerada", S2_L, bold=True)
lbl(ws, r, 2, "Informativo — FE 0 (reduce la compra de red)")
inp(ws, r, 3)
set_cell(ws, r, 4, "kWh/año", halign='center')
ref(ws, r, 5, 0)
calc(ws, r, 6, "=0")

r_sub = 5
merge(ws, r_sub, 1, r_sub, 5, "SUBTOTAL S2 — ELECTRICIDAD (tCO2e/año)", S2_D)
calc(ws, r_sub, 6, "=SUM(F3:F4)")
SUBTOT['S2 ELECTRICIDAD'] = f"F{r_sub}"

nota(ws, 6, 1, 6,
     "GHG Protocol Scope 2 Guidance (2015) — enfoque localización. FE Red Argentina CAMMESA 2023: "
     "0,383 kg CO2e/kWh. Verificar el valor vigente del año del inventario en "
     "cammesaweb.cammesa.com/download/factor-de-emision. Fuente del dato: facturas de la distribuidora (12 meses).")

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 8: S3 ENVASES (la fuente dominante en vino)
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("S3 ENVASES")
ws.sheet_properties.tabColor = S3_D
set_col_widths(ws, [28, 34, 16, 16, 16, 14, 14])

merge(ws, 1, 1, 1, 7, "SCOPE 3 — ENVASES Y EMBALAJES (fabricación)  |  Típicamente la mayor fuente en una bodega", S3_D, sz=13)

# Sección botellas, con % reciclado
merge(ws, 2, 1, 2, 7, "BOTELLAS DE VIDRIO", S3_D, sz=11)
for j, htxt in enumerate(["Concepto", "Descripción", "Dato", "Unidad",
                          "FE base (kg CO2e/kg)", "", "tCO2e/año"]):
    header(ws, 3, j + 1, htxt)

lbl(ws, 4, 1, "Botellas compradas en el año", S3_L, bold=True)
lbl(ws, 4, 2, "Unidades (todas las líneas y formatos)")
inp(ws, 4, 3, 240000)
set_cell(ws, 4, 4, "unidades/año", halign='center')
ref(ws, 4, 5, 1.2)
set_cell(ws, 4, 6, "")
set_cell(ws, 4, 7, "")

lbl(ws, 5, 1, "Peso promedio por botella", S3_L, bold=True)
lbl(ws, 5, 2, "Estándar 450-550 g; premium 700-900 g")
inp(ws, 5, 3, 500)
set_cell(ws, 5, 4, "gramos", halign='center')
set_cell(ws, 5, 5, "")
set_cell(ws, 5, 6, "")
set_cell(ws, 5, 7, "")

lbl(ws, 6, 1, "Contenido de vidrio reciclado", S3_L, bold=True)
lbl(ws, 6, 2, "Consultar al proveedor (botella oscura admite más casco)")
inp(ws, 6, 3, 30)
set_cell(ws, 6, 4, "%", halign='center')
set_cell(ws, 6, 5, "")
set_cell(ws, 6, 6, "")
set_cell(ws, 6, 7, "")

lbl(ws, 7, 1, "EMISIONES BOTELLAS", S3_L, bold=True)
lbl(ws, 7, 2, "= unid. x peso(kg) x FE x (1 - 0,002 x %reciclado)")
set_cell(ws, 7, 3, "")
set_cell(ws, 7, 4, "")
set_cell(ws, 7, 5, "")
set_cell(ws, 7, 6, "")
calc(ws, 7, 7, '=IF(C4="",0,C4*C5/1000*E4*(1-0.002*C6)/1000)')

# Sección tapones y resto del packaging
merge(ws, 8, 1, 8, 7, "TAPONES, CÁPSULAS, ETIQUETAS Y CAJAS", S3_D, sz=11)
for j, htxt in enumerate(["Ítem", "Descripción", "Dato anual", "Unidad",
                          "FE", "Unidad FE", "tCO2e/año"]):
    header(ws, 9, j + 1, htxt)

pack_rows = [
    ("Corcho natural",        "El cierre de menor huella",            "unidades/año", 0.00183, "kg CO2e/unidad"),
    ("Tapón sintético",       "Polímero",                             "unidades/año", 0.0148,  "kg CO2e/unidad"),
    ("Tapa rosca (aluminio)", "Screwcap",                             "unidades/año", 0.0372,  "kg CO2e/unidad"),
    ("Cápsulas (aluminio/estaño)", "Peso total comprado en el año",   "kg/año",       8.6,     "kg CO2e/kg"),
    ("Etiquetas (papel)",     "Peso total comprado en el año",        "kg/año",       1.1,     "kg CO2e/kg"),
    ("Cajas de cartón",       "Peso total (cajas + separadores)",     "kg/año",       0.94,    "kg CO2e/kg"),
]
for i, (item, desc, un, fe, unfe) in enumerate(pack_rows):
    r = 10 + i
    lbl(ws, r, 1, item, S3_L, bold=True)
    lbl(ws, r, 2, desc)
    inp(ws, r, 3)
    set_cell(ws, r, 4, un, halign='center')
    ref(ws, r, 5, fe)
    set_cell(ws, r, 6, unfe, halign='center')
    calc(ws, r, 7, f'=IF(C{r}="",0,C{r}*E{r}/1000)')

r_sub = 16
merge(ws, r_sub, 1, r_sub, 6, "SUBTOTAL S3 — ENVASES Y EMBALAJES (tCO2e/año)", S3_D)
calc(ws, r_sub, 7, "=G7+SUM(G10:G15)")
SUBTOT['S3 ENVASES'] = f"G{r_sub}"

nota(ws, 17, 1, 7,
     "Vidrio: 1,2 kg CO2e/kg (vidrio nuevo, promedio LCA europeos); cada 10% de casco reciclado reduce ~2% "
     "(Glass Alliance Europe). Cierres: corcho 1,83 g / sintético 14,8 g / rosca aluminio 37,2 g CO2e por unidad "
     "(estudio PwC para Amorim, 2008 — produccion). Aluminio 8,6 kg CO2e/kg (promedio primario). "
     "Cartón 0,94 kg CO2e/kg (cradle-to-grave). Papel 1,1 kg CO2e/kg. Ajustar con datos del proveedor si los entrega (EPD).")

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 9: S3 INSUMOS (fabricación, upstream)
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("S3 INSUMOS")
ws.sheet_properties.tabColor = S3_D
set_col_widths(ws, [30, 34, 16, 14, 16, 14, 14])

merge(ws, 1, 1, 1, 7, "SCOPE 3 — INSUMOS COMPRADOS (emisiones de fabricación, aguas arriba)", S3_D, sz=13)
for j, htxt in enumerate(["Insumo", "Descripción", "Dato anual", "Unidad",
                          "FE (kg CO2e/kg)", "Estado", "tCO2e/año"]):
    header(ws, 2, j + 1, htxt)

rows = [
    ("Fertilizantes (fabricación)", "Todo fertilizante comprado, por peso de producto", "kg/año", 1.2,  "Calculado"),
    ("Agroquímicos (fabricación)",  "Herbicidas, fungicidas, insecticidas",             "kg o L/año", 6.3, "Calculado"),
    ("Productos enológicos",        "Levaduras, clarificantes, SO2, ácidos — sin FE estandarizado", "kg/año", None, "Registro"),
]
for i, (item, desc, un, fe, estado) in enumerate(rows):
    r = 3 + i
    lbl(ws, r, 1, item, S3_L, bold=True)
    lbl(ws, r, 2, desc)
    inp(ws, r, 3)
    set_cell(ws, r, 4, un, halign='center')
    if fe is not None:
        ref(ws, r, 5, fe)
        set_cell(ws, r, 6, estado, halign='center')
        calc(ws, r, 7, f'=IF(C{r}="",0,C{r}*E{r}/1000)')
    else:
        set_cell(ws, r, 5, "—", halign='center')
        set_cell(ws, r, 6, estado, halign='center')
        calc(ws, r, 7, "=0")

r_sub = 7
merge(ws, r_sub, 1, r_sub, 6, "SUBTOTAL S3 — INSUMOS (tCO2e/año)", S3_D)
calc(ws, r_sub, 7, "=SUM(G3:G5)")
SUBTOT['S3 INSUMOS'] = f"G{r_sub}"

nota(ws, 8, 1, 7,
     "FE promedio de bibliografía (fabricación): fertilizantes 1,2 / agroquímicos 6,3 kg CO2e/kg. "
     "Los productos enológicos se registran sin calcular hasta contar con FE específicos o EPD del proveedor.")

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 10: S3 TRANSPORTE Y DISTRIBUCIÓN
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("S3 TRANSPORTE")
ws.sheet_properties.tabColor = S3_D
set_col_widths(ws, [30, 34, 16, 14, 18, 14, 14])

merge(ws, 1, 1, 1, 7, "SCOPE 3 — TRANSPORTE Y DISTRIBUCIÓN (contratado a terceros)", S3_D, sz=13)
for j, htxt in enumerate(["Tramo", "Descripción", "Dato anual", "Unidad",
                          "FE (kg CO2e/t·km)", "Modo", "tCO2e/año"]):
    header(ws, 2, j + 1, htxt)

rows = [
    ("Insumos hacia la bodega",   "Botellas, insumos secos, fertilizantes (t x km)", 0.062, "Camión"),
    ("Distribución nacional",     "Bodega -> distribuidores / clientes (t x km)",    0.062, "Camión"),
    ("Tramo terrestre a puerto",  "Bodega -> puerto de salida (t x km)",             0.062, "Camión"),
    ("Flete marítimo exportación","Puerto -> destino (t x km náuticos convertidos)", 0.012, "Buque contenedor"),
    ("Ferrocarril (si aplica)",   "Tramos ferroviarios de carga (t x km)",           0.022, "Tren"),
]
for i, (item, desc, fe, modo) in enumerate(rows):
    r = 3 + i
    lbl(ws, r, 1, item, S3_L, bold=True)
    lbl(ws, r, 2, desc)
    inp(ws, r, 3)
    set_cell(ws, r, 4, "t·km/año", halign='center')
    ref(ws, r, 5, fe)
    set_cell(ws, r, 6, modo, halign='center')
    calc(ws, r, 7, f'=IF(C{r}="",0,C{r}*E{r}/1000)')

r_sub = 8
merge(ws, r_sub, 1, r_sub, 6, "SUBTOTAL S3 — TRANSPORTE (tCO2e/año)", S3_D)
calc(ws, r_sub, 7, "=SUM(G3:G7)")
SUBTOT['S3 TRANSPORTE'] = f"G{r_sub}"

nota(ws, 9, 1, 7,
     "t·km = toneladas transportadas x kilómetros recorridos. Ej.: 120 t de vino a un puerto a 1.200 km = "
     "144.000 t·km. FE de referencia: camión carga completa 0,062 / tren 0,022 / buque contenedor 0,012 "
     "kg CO2e/t·km (valores medios bibliografía logística, ajustables con datos del operador).")

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 11: FACTORES DE EMISIÓN (solo lectura)
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("FACTORES EMISION")
ws.sheet_properties.tabColor = NAVY
set_col_widths(ws, [36, 34, 14, 16, 40])

merge(ws, 1, 1, 1, 5, "TABLA DE FACTORES DE EMISIÓN DE REFERENCIA — SOLO LECTURA", NAVY, sz=13)
nota(ws, 2, 1, 5,
     "Fuentes: IPCC 2006 / 2019 Refinement | GWP IPCC AR6 (2021) | GHG Protocol | CAMMESA | "
     "Protocolo OIV GEI | PwC-Amorim (cierres) | Glass Alliance Europe (vidrio). NO modificar esta hoja.")

for j, htxt in enumerate(["Fuente / Actividad", "Aplicación", "FE", "Unidad", "Fuente bibliográfica"]):
    header(ws, 3, j + 1, htxt)

factores = [
    ("Gasoil (diesel)",            "Combustión móvil y fija",   2.68,    "kg CO2e/L",      "IPCC 2006 Vol. 2 — CO2+CH4+N2O combinado"),
    ("Nafta (gasolina)",           "Vehículos livianos",        2.31,    "kg CO2e/L",      "IPCC 2006 Vol. 2"),
    ("Gas natural",                "Caldera / calefacción",     2.04,    "kg CO2e/m3",     "IPCC 2006 Vol. 2"),
    ("GLP",                        "Garrafas / zeppelin",       3.02,    "kg CO2e/kg",     "IPCC 2006 Vol. 2"),
    ("Nitrógeno al suelo (N2O directo)", "Fertilización del viñedo", 4.29, "kg CO2e/kg N", "IPCC 2019 Ref. EF1=0,01 x 44/28 x GWP 273 (AR6)"),
    ("Urea — hidrólisis",          "Aplicación de urea",        0.733,   "kg CO2/kg urea", "IPCC 2006 Vol. 4 (0,20 C x 44/12)"),
    ("Cal agrícola",               "Encalado",                  0.440,   "kg CO2/kg",      "IPCC 2006 Vol. 4"),
    ("Dolomita",                   "Encalado",                  0.477,   "kg CO2/kg",      "IPCC 2006 Vol. 4"),
    ("R-134a",                     "Refrigerante chillers",     1530,    "GWP 100 años",   "IPCC AR6 (2021)"),
    ("R-404A",                     "Refrigerante cámaras",      4728,    "GWP 100 años",   "IPCC AR6 (2021) — mezcla"),
    ("R-410A",                     "Aire acondicionado",        2256,    "GWP 100 años",   "IPCC AR6 (2021) — mezcla"),
    ("R-22 (HCFC)",                "Equipos antiguos",          1960,    "GWP 100 años",   "IPCC AR6 (2021)"),
    ("Red eléctrica Argentina",    "Scope 2 — localización",    0.383,   "kg CO2e/kWh",    "CAMMESA 2023 — verificar vigencia anual"),
    ("Vidrio nuevo (botellas)",    "Fabricación de envase",     1.2,     "kg CO2e/kg",     "Promedio LCA europeos; -2% por cada 10% reciclado (Glass Alliance)"),
    ("Corcho natural",             "Cierre",                    0.00183, "kg CO2e/unidad", "PwC para Amorim (2008) — fase producción"),
    ("Tapón sintético",            "Cierre",                    0.0148,  "kg CO2e/unidad", "PwC para Amorim (2008)"),
    ("Tapa rosca aluminio",        "Cierre",                    0.0372,  "kg CO2e/unidad", "PwC para Amorim (2008)"),
    ("Aluminio (cápsulas)",        "Packaging",                 8.6,     "kg CO2e/kg",     "Promedio aluminio primario — bibliografía"),
    ("Papel (etiquetas)",          "Packaging",                 1.1,     "kg CO2e/kg",     "Promedio bibliografía packaging"),
    ("Cartón (cajas)",             "Packaging",                 0.94,    "kg CO2e/kg",     "LCA cradle-to-grave cartón corrugado"),
    ("Fertilizantes (fabricación)","Scope 3 upstream",          1.2,     "kg CO2e/kg",     "Promedio bibliografía"),
    ("Agroquímicos (fabricación)", "Scope 3 upstream",          6.3,     "kg CO2e/kg",     "Promedio bibliografía"),
    ("Camión de carga",            "Transporte",                0.062,   "kg CO2e/t·km",   "Valores medios logística"),
    ("Tren de carga",              "Transporte",                0.022,   "kg CO2e/t·km",   "Valores medios logística"),
    ("Buque portacontenedores",    "Exportación marítima",      0.012,   "kg CO2e/t·km",   "Valores medios logística"),
    ("CO2 fermentación (biogénico)","Informativo — fuera del total", 0.098, "kg CO2/L vino", "Estequiometría, vino 13% vol — criterio OIV/FIVS"),
]
for i, (item, ap, fe, un, fuente) in enumerate(factores):
    r = 4 + i
    lbl(ws, r, 1, item, GR1, bold=True)
    lbl(ws, r, 2, ap)
    ref(ws, r, 3, fe)
    set_cell(ws, r, 4, un, halign='center')
    lbl(ws, r, 5, fuente)

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 12: RESUMEN
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("RESUMEN")
ws.sheet_properties.tabColor = OR_D
set_col_widths(ws, [14, 44, 14, 16, 44])

merge(ws, 1, 1, 1, 5, "RESUMEN DEL INVENTARIO — HUELLA DE CARBONO DE LA BODEGA", OR_D, sz=14)
for j, htxt in enumerate(["Código", "Categoría", "Alcance", "tCO2e/año", "Origen del dato"]):
    header(ws, 2, j + 1, htxt)

resumen = [
    ("S1-1", "Combustión móvil (maquinaria y vehículos)", "Scope 1", "'S1 MAQUINARIA'!"   + SUBTOT['S1 MAQUINARIA'],   S1_L),
    ("S1-2", "Combustión estacionaria",                   "Scope 1", "'S1 ENERGIA FIJA'!" + SUBTOT['S1 ENERGIA FIJA'], S1_L),
    ("S1-3", "Viñedo — N2O y enmiendas",                  "Scope 1", "'S1 VIÑEDO'!"       + SUBTOT['S1 VIÑEDO'],       S1_L),
    ("S1-4", "Refrigerantes (fugas)",                     "Scope 1", "'S1 REFRIGERANTES'!"+ SUBTOT['S1 REFRIGERANTES'],S1_L),
    ("S2-1", "Electricidad de red",                       "Scope 2", "'S2 ELECTRICIDAD'!" + SUBTOT['S2 ELECTRICIDAD'], S2_L),
    ("S3-1", "Envases y embalajes",                       "Scope 3", "'S3 ENVASES'!"      + SUBTOT['S3 ENVASES'],      S3_L),
    ("S3-2", "Insumos comprados (fabricación)",           "Scope 3", "'S3 INSUMOS'!"      + SUBTOT['S3 INSUMOS'],      S3_L),
    ("S3-3", "Transporte y distribución",                 "Scope 3", "'S3 TRANSPORTE'!"   + SUBTOT['S3 TRANSPORTE'],   S3_L),
]
for i, (code, cat, scope, ref_cell, bg) in enumerate(resumen):
    r = 3 + i
    set_cell(ws, r, 1, code, bg=bg, bold=True, halign='center')
    lbl(ws, r, 2, cat, bg, bold=True)
    set_cell(ws, r, 3, scope, bg=bg, halign='center')
    calc(ws, r, 4, "=" + ref_cell)
    lbl(ws, r, 5, "Subtotal automático de la hoja " + ref_cell.split('!')[0].strip("'"))

# Totales por alcance
merge(ws, 12, 1, 12, 3, "TOTAL SCOPE 1 (tCO2e/año)", S1_D)
calc(ws, 12, 4, "=SUM(D3:D6)")
set_cell(ws, 12, 5, "")
merge(ws, 13, 1, 13, 3, "TOTAL SCOPE 2 (tCO2e/año)", S2_D)
calc(ws, 13, 4, "=D7")
set_cell(ws, 13, 5, "")
merge(ws, 14, 1, 14, 3, "TOTAL SCOPE 3 (tCO2e/año)", S3_D)
calc(ws, 14, 4, "=SUM(D8:D10)")
set_cell(ws, 14, 5, "")
merge(ws, 15, 1, 15, 3, "TOTAL HUELLA DE CARBONO (S1+S2+S3)", OR_D, sz=12)
calc(ws, 15, 4, "=D12+D13+D14")
set_cell(ws, 15, 5, "")

# Biogénico aparte
merge(ws, 17, 1, 17, 3, "CO2 biogénico de fermentación (informativo — NO suma)", BIO_D)
calc(ws, 17, 4, "='FERMENTACION (INFO)'!" + SUBTOT['FERMENTACION (INFO)'])
lbl(ws, 17, 5, "Criterio OIV/FIVS: carbono de ciclo corto, se reporta por separado")

# Intensidades
merge(ws, 19, 1, 19, 5, "INTENSIDAD DE EMISIONES (indicador organizacional)", NAVY, sz=11)
lbl(ws, 20, 1, "", GR1); lbl(ws, 20, 2, "Botellas producidas por año (de PORTADA)", GR1, bold=True)
set_cell(ws, 20, 3, "", bg=GR1)
calc(ws, 20, 4, "=PORTADA!B11")
lbl(ws, 20, 5, "Equivalente 0,75 L")
lbl(ws, 21, 1, "", GR1); lbl(ws, 21, 2, "kg CO2e por botella (S1+S2+S3)", GR1, bold=True)
set_cell(ws, 21, 3, "", bg=GR1)
calc(ws, 21, 4, "=IF(D20=0,0,D15*1000/D20)")
lbl(ws, 21, 5, "Benchmark bibliográfico: 0,9 a 1,9 kg CO2e/botella (cradle-to-grave)")
lbl(ws, 22, 1, "", GR1); lbl(ws, 22, 2, "kg CO2e por litro de vino", GR1, bold=True)
set_cell(ws, 22, 3, "", bg=GR1)
calc(ws, 22, 4, "=IF(PORTADA!B10=0,0,D15*1000/PORTADA!B10)")
lbl(ws, 22, 5, "Litros de PORTADA")

nota(ws, 24, 1, 5,
     "La intensidad por botella es un indicador ORGANIZACIONAL (total de la bodega dividido por la producción). "
     "No constituye una huella de producto certificable bajo ISO 14067, que exige asignación por línea de producto "
     "y ciclo de vida completo. Es válida y útil como indicador de gestión y para comunicación con compradores, "
     "declarando la metodología.")

# ─────────────────────────────────────────────────────────────────────────────
import os
out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "Matriz_HuellaCarbono_Vitivinicola.xlsx")
wb.save(out)
print(f"OK -> {out}")
print(f"Hojas: {wb.sheetnames}")
