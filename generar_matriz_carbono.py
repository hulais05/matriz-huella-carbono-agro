#!/usr/bin/env python3
"""
Genera la Matriz de Recopilación de Datos — Huella de Carbono
Sector Agropecuario | GHG Protocol + ISO 14064-1
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def brd(style='thin'):
    s = Side(style=style, color='FF000000')
    return Border(left=s, right=s, top=s, bottom=s)

def fl(hex6):
    return PatternFill(fill_type='solid', fgColor='FF' + hex6)

def set_cell(ws, r, c, value=None, bg='F2F2F2', fg='000000',
             bold=False, sz=10, halign='left', valign='center',
             wrap=True, italic=False, border_style='thin'):
    cell = ws.cell(row=r, column=c, value=value)
    cell.fill = fl(bg)
    cell.font = Font(bold=bold, color='FF' + fg, size=sz, italic=italic)
    cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    cell.border = brd(border_style)
    return cell

def header(ws, r, c, text, bg='1F3864', fg='FFFFFF', sz=10, bold=True, halign='center'):
    return set_cell(ws, r, c, text, bg=bg, fg=fg, bold=bold, sz=sz, halign=halign)

def inp(ws, r, c, value=None):
    """Celda de ingreso — amarillo"""
    return set_cell(ws, r, c, value, bg='FFFF99', fg='000000', halign='center')

def calc(ws, r, c, formula):
    """Celda calculada — verde"""
    return set_cell(ws, r, c, formula, bg='E2EFDA', fg='375623', italic=True, halign='center')

def ref(ws, r, c, value):
    """Factor de referencia — celeste"""
    return set_cell(ws, r, c, value, bg='BDD7EE', fg='1F3864', halign='center')

def lbl(ws, r, c, text, bg='F2F2F2', bold=False):
    return set_cell(ws, r, c, text, bg=bg, fg='000000', bold=bold, halign='left')

def merge(ws, r1, c1, r2, c2, text, bg='1F3864', fg='FFFFFF',
          bold=True, sz=11, halign='center'):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=text)
    cell.fill = fl(bg)
    cell.font = Font(bold=bold, color='FF' + fg, size=sz)
    cell.alignment = Alignment(horizontal=halign, vertical='center', wrap_text=True)
    cell.border = brd()
    return cell

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def set_row_h(ws, rows_heights):
    for r, h in rows_heights.items():
        ws.row_dimensions[r].height = h

# ─── COLORES ─────────────────────────────────────────────────────────────────
NAVY    = '1F3864'
TITLE   = '243F60'
S1_D    = 'C00000'   # Scope 1 oscuro
S1_L    = 'FFC7CE'   # Scope 1 claro
S2_D    = '0070C0'   # Scope 2 oscuro
S2_L    = 'DCE6F1'   # Scope 2 claro
SK_D    = '375623'   # Sumideros oscuro
SK_L    = 'C6EFCE'   # Sumideros claro
OR_D    = 'ED7D31'   # Resumen oscuro
OR_L    = 'FCE4D6'   # Resumen claro
GR1     = 'F2F2F2'
GR2     = 'D9D9D9'
GR3     = 'BFBFBF'

wb = openpyxl.Workbook()

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 1: PORTADA
# ─────────────────────────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "PORTADA"
ws1.sheet_properties.tabColor = TITLE

set_row_h(ws1, {1: 45, 2: 28, 3: 12, 4: 28, 5: 22, 6: 22, 7: 22, 8: 22,
                9: 22, 10: 22, 11: 22, 12: 22, 13: 22, 14: 22, 15: 22,
                16: 22, 17: 12, 18: 28, 19: 22, 20: 22, 21: 22, 22: 22,
                23: 22, 24: 22, 25: 22, 26: 22, 27: 22, 28: 22, 29: 22})

merge(ws1, 1, 1, 1, 9, "MATRIZ DE RECOPILACIÓN DE DATOS — HUELLA DE CARBONO CORPORATIVA",
      TITLE, sz=16)
merge(ws1, 2, 1, 2, 9, "GHG Protocol (Corporate Standard)  +  ISO 14064-1:2018  |  Sector Agropecuario",
      TITLE, sz=11)

# Identificación
merge(ws1, 4, 1, 4, 5, "IDENTIFICACIÓN DEL ESTABLECIMIENTO", NAVY, sz=11)
id_fields = [
    ("Razón Social / Nombre del Establecimiento",),
    ("CUIT / RUT",),
    ("Provincia / Departamento",),
    ("Dirección del establecimiento",),
    ("Superficie total declarada (ha)", "1.000"),
    ("Superficie productiva (ha)", "900"),
    ("Año del inventario",),
    ("Año base del inventario",),
    ("Responsable del inventario",),
    ("Cargo / Matrícula profesional",),
    ("Correo electrónico",),
    ("Teléfono de contacto",),
]
for i, row_data in enumerate(id_fields):
    r = 5 + i
    lbl(ws1, r, 1, row_data[0], GR1, bold=True)
    ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    val = row_data[1] if len(row_data) > 1 else None
    inp(ws1, r, 2, val)
    for c in range(3, 6):
        cell = ws1.cell(row=r, column=c)
        cell.fill = fl('FFFF99')
        cell.border = brd()

# Límites del inventario
merge(ws1, 4, 6, 4, 9, "LÍMITES DEL INVENTARIO", NAVY, sz=11)
lim_fields = [
    ("Enfoque organizacional", "Control Operacional"),
    ("Alcance cubierto", "Scope 1 + Scope 2"),
    ("¿Incluye sumideros?", "Sí"),
    ("FE maquinaria/combustión", "IPCC 2006 / GHG Protocol"),
    ("FE ganadería", "IPCC 2006 Tier 1/2"),
    ("FE electricidad", "Factor Red Argentina (CAMMESA)"),
    ("Factor red eléctrica (kg CO2e/kWh)", "0,383 (CAMMESA 2023 — verificar valor vigente cada año)"),
    ("PRG CH4 fósil — 100 años (AR6)", "29,8 (combustión gasoil/GLP)"),
    ("PRG CH4 biogénico — 100 años (AR6)", "27,2 (ganadería/efluentes)"),
    ("PRG N2O — 100 años (AR6)", "273"),
    ("Unidad de reporte", "tCO2e / año"),
    ("Versión del protocolo GHG", "Corporate Standard Rev. 2015"),
    ("Norma de referencia", "ISO 14064-1:2018"),
]
for i, (field, default) in enumerate(lim_fields):
    r = 5 + i
    lbl(ws1, r, 6, field, S2_L, bold=True)
    ws1.merge_cells(start_row=r, start_column=7, end_row=r, end_column=9)
    ref(ws1, r, 7, default)
    for c in range(8, 10):
        cell = ws1.cell(row=r, column=c)
        cell.fill = fl('BDD7EE')
        cell.border = brd()

# Leyenda
merge(ws1, 18, 1, 18, 9, "LEYENDA DE COLORES", GR2, '000000', sz=10)
legend_items = [
    (1, 2, "Celda de INGRESO DE DATOS (completar)", 'FFFF99', '000000'),
    (3, 4, "Celda CALCULADA automáticamente", 'E2EFDA', '375623'),
    (5, 6, "Factor de Emisión de referencia", 'BDD7EE', '1F3864'),
    (7, 8, "Encabezado / Categoría", GR2, '000000'),
]
for c1, c2, text, bg, fg in legend_items:
    ws1.merge_cells(start_row=19, start_column=c1, end_row=19, end_column=c2)
    cell = ws1.cell(row=19, column=c1, value=text)
    cell.fill = fl(bg)
    cell.font = Font(bold=True, color='FF' + fg, size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = brd()

# Actividades declaradas
merge(ws1, 21, 1, 21, 9, "FUENTES DE EMISIÓN Y SUMIDEROS DECLARADOS", NAVY, sz=11)
actividades = [
    ("S1-1", "Combustión Móvil", "Tractores, cosechadoras, camiones — gasoil"),
    ("S1-2", "Combustión Estacionaria", "Generadores a gasoil + Zeppelin de GLP"),
    ("S1-3", "Fermentación Entérica", "Bovinos feedlot y tambo — CH4 ruminal"),
    ("S1-4", "Gestión de Estiércol", "CH4 y N2O de residuos feedlot y tambo"),
    ("S1-5", "Suelos Agrícolas", "N2O de fertilizantes nitrogenados y urea"),
    ("S1-6", "Efluentes", "CH4/N2O de lagunas feedlot y tambo"),
    ("S2-1", "Electricidad de Red", "20% del consumo eléctrico total"),
    ("SK-1", "Sumideros", "Monte nativo y pasturas — capturas de CO2"),
]
hdr_cols_act = ["Código", "Fuente / Sumidero", "Descripción"]
for j, h_text in enumerate(hdr_cols_act):
    header(ws1, 22, j + 1 if j < 2 else j * 3 - 1, h_text,
           bg=GR3, fg='000000', sz=9)

for i, (code, name, desc) in enumerate(actividades):
    r = 23 + i
    ws1.row_dimensions[r].height = 20
    scope_bg = S1_L if code.startswith('S1') else (S2_L if code.startswith('S2') else SK_L)
    set_cell(ws1, r, 1, code, bg=scope_bg, fg='000000', bold=True, halign='center')
    set_cell(ws1, r, 2, name, bg=scope_bg, fg='000000', bold=True)
    ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)

set_col_widths(ws1, [32, 14, 14, 14, 14, 32, 14, 14, 14])

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 2: S1 — COMBUSTIÓN MÓVIL
# ─────────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("S1-1 MAQUINARIA")
ws2.sheet_properties.tabColor = S1_D
ws2.freeze_panes = 'A5'

merge(ws2, 1, 1, 1, 14, "SCOPE 1 — COMBUSTIÓN MÓVIL  |  Tractores, Cosechadoras y Camiones", S1_D, sz=13)
merge(ws2, 2, 1, 2, 14,
      "GHG Protocol — Mobile Combustion  |  IPCC 2006 Vol.2 Cap.3  |  FE gasoil: 2,7 kg CO2/L  |  CH4: 0,1 g/MJ  |  N2O: 0,6 g/MJ",
      S1_L, '000000', bold=False, sz=9)

cols_maq = [
    "N°", "Tipo de Máquina", "Marca / Modelo", "Año\nFabric.", "Uso\nAnual\n(meses/año)",
    "Horas de\nUso / mes\n[INGRESAR]", "Consumo\npromedio\n(L/hora)\n[INGRESAR]",
    "Consumo\nTotal\nAnual (L)\n[CALC]",
    "FE CO2\n(kg CO2/L)", "FE CH4\n(g/L)", "FE N2O\n(g/L)",
    "CO2\n(kg/año)", "CH4\n(kg CO2e/año)", "N2O\n(kg CO2e/año)", "TOTAL\ntCO2e/año"
]
for j, text in enumerate(cols_maq):
    header(ws2, 4, j + 1, text, bg=S1_D, sz=9)

maquinas = [
    ("1", "Tractor", "", "", "12"),
    ("2", "Tractor", "", "", "12"),
    ("3", "Cosechadora", "", "", "3"),
    ("4", "Cosechadora", "", "", "3"),
    ("5", "Cosechadora", "", "", "3"),
    ("6", "Cosechadora", "", "", "3"),
    ("7", "Cosechadora", "", "", "3"),
    ("8", "Camión", "", "", "12"),
    ("9", "Camión", "", "", "12"),
]

for i, (num, tipo, marca, anio, meses) in enumerate(maquinas):
    r = 5 + i
    ws2.row_dimensions[r].height = 20
    bg_row = 'FFFFFF' if i % 2 == 0 else 'F9F9F9'
    set_cell(ws2, r, 1, num, bg=bg_row, halign='center')
    lbl(ws2, r, 2, tipo, bg=bg_row)
    inp(ws2, r, 3, marca if marca else None)   # Marca
    inp(ws2, r, 4, None)                        # Año
    inp(ws2, r, 5, meses)                       # Meses uso
    inp(ws2, r, 6, None)                        # Horas/mes
    inp(ws2, r, 7, None)                        # L/hora
    # Consumo total = horas/mes * meses * L/hora
    calc(ws2, r, 8, f"=F{r}*E{r}*G{r}")
    ref(ws2, r, 9, 2.701)    # FE CO2 gasoil kg CO2/L (IPCC)
    ref(ws2, r, 10, 0.100)   # FE CH4 g/L aproximado
    ref(ws2, r, 11, 0.600)   # FE N2O g/L aproximado
    calc(ws2, r, 12, f"=H{r}*I{r}")
    calc(ws2, r, 13, f"=H{r}*J{r}/1000*29.8")
    calc(ws2, r, 14, f"=H{r}*K{r}/1000*273")
    calc(ws2, r, 15, f"=(L{r}+M{r}+N{r})/1000")

# Fila TOTAL
r_tot = 14
ws2.row_dimensions[r_tot].height = 22
merge(ws2, r_tot, 1, r_tot, 11, "TOTAL COMBUSTIÓN MÓVIL", GR2, '000000', bold=True, sz=10)
calc(ws2, r_tot, 12, f"=SUM(L5:L{r_tot-1})")
calc(ws2, r_tot, 13, f"=SUM(M5:M{r_tot-1})")
calc(ws2, r_tot, 14, f"=SUM(N5:N{r_tot-1})")
calc(ws2, r_tot, 15, f"=SUM(O5:O{r_tot-1})")

# Nota
r_n = r_tot + 2
merge(ws2, r_n, 1, r_n, 15,
      "NOTA: Para camiones con datos en km/año en lugar de horas, usar consumo en L/km y reemplazar columnas E y F por km/mes y L/km. "
      "FE gasoil = 2,701 kg CO2/L (IPCC 2006 Tabla 1.4). "
      "Si se usa biodiesel B5/B10, ajustar FE. Fuente de datos sugerida: planilla de control de combustible, "
      "remitos de surtidor o estimación por horas de horómetro.",
      'FFF2CC', '7F6000', bold=False, sz=9)

set_col_widths(ws2, [4, 14, 16, 7, 8, 11, 11, 11, 8, 7, 7, 11, 13, 13, 11])

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 3: S1 — ENERGÍA ESTACIONARIA
# ─────────────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("S1-2 ENERGIA ESTAC")
ws3.sheet_properties.tabColor = S1_D
ws3.freeze_panes = 'A5'

merge(ws3, 1, 1, 1, 12, "SCOPE 1 — COMBUSTIÓN ESTACIONARIA  |  Generadores a Gasoil + Zeppelin GLP", S1_D, sz=13)
merge(ws3, 2, 1, 2, 12,
      "GHG Protocol — Stationary Combustion  |  IPCC 2006 Vol.2 Cap.2  |  FE gasoil estacionario: 2,701 kg CO2/L  |  FE GLP: 2,983 kg CO2/kg",
      S1_L, '000000', bold=False, sz=9)

# Sección A — Generadores
merge(ws3, 3, 1, 3, 12, "SECCIÓN A — GENERADORES ELÉCTRICOS (Autogeneración a gasoil)", S1_D, sz=10)
cols_gen = ["N°", "Descripción / Ubicación", "Potencia\n(kW)", "Meses\nOper.\n/año",
            "Horas\nFuncion.\n/mes\n[INGRESAR]", "Consumo\npromedio\n(L/hora)\n[INGRESAR]",
            "Consumo\nTotal\nAnual (L)\n[CALC]", "kWh\nGenerados\n/año\n[INGRESAR]",
            "FE CO2\n(kg/L)", "CO2\n(kg/año)", "CH4\n(kg CO2e\n/año)", "TOTAL\ntCO2e/año"]
for j, text in enumerate(cols_gen):
    header(ws3, 4, j + 1, text, bg=S1_D, sz=9)

gens = [("1", "Generador principal"), ("2", "Generador respaldo")]
for i, (num, desc) in enumerate(gens):
    r = 5 + i
    ws3.row_dimensions[r].height = 20
    bg_r = 'FFFFFF' if i % 2 == 0 else 'F9F9F9'
    set_cell(ws3, r, 1, num, bg=bg_r, halign='center')
    lbl(ws3, r, 2, desc, bg=bg_r)
    inp(ws3, r, 3, None)
    inp(ws3, r, 4, "12")
    inp(ws3, r, 5, None)
    inp(ws3, r, 6, None)
    calc(ws3, r, 7, f"=E{r}*D{r}*F{r}")
    inp(ws3, r, 8, None)
    ref(ws3, r, 9, 2.701)
    calc(ws3, r, 10, f"=G{r}*I{r}")
    calc(ws3, r, 11, f"=G{r}*0.0001*29.8")
    calc(ws3, r, 12, f"=(J{r}+K{r})/1000")

r_tot_a = 7
merge(ws3, r_tot_a, 1, r_tot_a, 9, "SUBTOTAL GENERADORES", GR2, '000000', bold=True, sz=10)
calc(ws3, r_tot_a, 10, "=SUM(J5:J6)")
calc(ws3, r_tot_a, 11, "=SUM(K5:K6)")
calc(ws3, r_tot_a, 12, "=SUM(L5:L6)")

# Sección B — Zeppelin GLP
merge(ws3, 9, 1, 9, 12, "SECCIÓN B — ZEPPELIN DE GAS LICUADO DE PETRÓLEO (GLP / Propano-Butano)", S1_D, sz=10)
cols_zlp = ["N°", "Descripción / Ubicación", "Capacidad\nZeppelin\n(kg)", "Recargas\npor año\n[INGRESAR]",
            "kg GLP\npor recarga\n[INGRESAR]", "Consumo\nTotal\nAnual (kg)\n[CALC]",
            "Litros\nequiv.\n/año\n[CALC]", "FE CO2\n(kg CO2/kg\nGLP)",
            "CO2\n(kg/año)", "CH4\n(kg CO2e\n/año)", "N2O\n(kg CO2e\n/año)", "TOTAL\ntCO2e/año"]
for j, text in enumerate(cols_zlp):
    header(ws3, 10, j + 1, text, bg=S1_D, sz=9)

ws3.row_dimensions[11].height = 20
set_cell(ws3, 11, 1, "1", bg='FFFFFF', halign='center')
lbl(ws3, 11, 2, "Zeppelin GLP - instalaciones")
inp(ws3, 11, 3, None)
inp(ws3, 11, 4, None)
inp(ws3, 11, 5, None)
calc(ws3, 11, 6, "=D11*E11")
calc(ws3, 11, 7, "=F11/0.51")   # densidad GLP aprox 0.51 kg/L
ref(ws3, 11, 8, 2.983)
calc(ws3, 11, 9, "=F11*H11")
calc(ws3, 11, 10, "=F11*0.001*29.8")
calc(ws3, 11, 11, "=F11*0.0001*273")
calc(ws3, 11, 12, "=(I11+J11+K11)/1000")

r_tot_b = 12
merge(ws3, r_tot_b, 1, r_tot_b, 8, "SUBTOTAL ZEPPELIN GLP", GR2, '000000', bold=True, sz=10)
calc(ws3, r_tot_b, 9, "=I11")
calc(ws3, r_tot_b, 10, "=J11")
calc(ws3, r_tot_b, 11, "=K11")
calc(ws3, r_tot_b, 12, "=L11")

# Total general
r_tot_c = 14
merge(ws3, r_tot_c, 1, r_tot_c, 9,
      "TOTAL COMBUSTIÓN ESTACIONARIA (Generadores + Zeppelin)", GR2, '000000', bold=True, sz=10)
calc(ws3, r_tot_c, 10, f"=J{r_tot_a}+I{r_tot_b}")
calc(ws3, r_tot_c, 11, f"=K{r_tot_a}+J{r_tot_b}")
calc(ws3, r_tot_c, 12, f"=L{r_tot_a}+L{r_tot_b}")

merge(ws3, 16, 1, 16, 12,
      "NOTA: FE GLP (propano comercial) = 2,983 kg CO2/kg (IPCC 2006). "
      "Para determinar el consumo real, registrar las remisiones/facturas de recarga del zeppelin. "
      "Si no se cuenta con datos previos, instalar un horómetro en generadores y registrar "
      "litros en cada carga de combustible desde el inicio del período de reporte.",
      'FFF2CC', '7F6000', bold=False, sz=9)

set_col_widths(ws3, [4, 22, 10, 10, 11, 11, 11, 10, 10, 11, 11, 11])

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 4: S1 — GANADERÍA (FERMENTACIÓN + ESTIÉRCOL)
# ─────────────────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet("S1-3 GANADERIA")
ws4.sheet_properties.tabColor = S1_D
ws4.freeze_panes = 'A6'

merge(ws4, 1, 1, 1, 14, "SCOPE 1 — GANADERÍA  |  Fermentación Entérica + Gestión de Estiércol", S1_D, sz=13)
merge(ws4, 2, 1, 2, 14,
      "IPCC 2006 Vol.4 Cap.10/11  |  Tier 1 por defecto — upgradeble a Tier 2 con datos de peso y dieta  |  "
      "PRG CH4 biogénico = 27,2  |  PRG N2O = 273  (AR6, 100 años)",
      S1_L, '000000', bold=False, sz=9)

# Sección A — Fermentación Entérica
merge(ws4, 3, 1, 3, 14, "SECCIÓN A — FERMENTACIÓN ENTÉRICA (CH4 ruminal)", S1_D, sz=10)
cols_ent = [
    "Categoría Animal", "Sistema\nProductivo",
    "Cabezas\ntotales\n[INGRESAR]", "Días en\nestab.\n/año\n[INGRESAR]",
    "Cabezas\nequiv.\n(ajustado)\n[CALC]",
    "Peso vivo\npromedio\n(kg)\n[INGRESAR]",
    "FE CH4\nFerm.Ent.\n(kg CH4\n/cab/año)\n[Tier 1]",
    "CH4\ntotal\n(kg/año)\n[CALC]",
    "CH4\n(kg CO2e\n/año)\n[CALC]",
    "Producción\nanual\nleche\no carne\n[INGRESAR]",
    "Unidad", "Observaciones"
]
for j, text in enumerate(cols_ent):
    header(ws4, 5, j + 1, text, bg=S1_D, sz=8)

categorias_gan = [
    ("Vacas en ordeñe (tambo)", "Tambo", None, "365", None, None, 119.8),
    ("Vaquillonas en ordeñe", "Tambo", None, "365", None, None, 100.0),
    ("Novillos feedlot", "Feedlot", None, None, None, None, 56.5),
    ("Novillitos feedlot", "Feedlot", None, None, None, None, 47.8),
    ("Vaquillonas de recría", "Cría/Recría", None, "365", None, None, 47.8),
    ("Terneros/as (dest.-12 meses)", "Cría", None, "365", None, None, 27.0),
    ("Toros", "Cría/Tambo", None, "365", None, None, 73.0),
    ("Ovinos / Caprinos", "Extensivo", None, "365", None, None, 8.0),
    ("Cerdos (si aplica)", "Confinado", None, "365", None, None, 1.5),
    ("Aves (si aplica)", "Confinado", None, "365", None, None, 0.02),
]
for i, (cat, sist, cabs, dias, ceq, pv, fe) in enumerate(categorias_gan):
    r = 6 + i
    ws4.row_dimensions[r].height = 22
    bg_r = 'FFFFFF' if i % 2 == 0 else 'F9F9F9'
    lbl(ws4, r, 1, cat, bg_r, bold=True)
    lbl(ws4, r, 2, sist, bg_r)
    inp(ws4, r, 3, cabs)
    inp(ws4, r, 4, dias)
    calc(ws4, r, 5, f"=C{r}*D{r}/365")
    inp(ws4, r, 6, pv)
    ref(ws4, r, 7, fe)
    calc(ws4, r, 8, f"=E{r}*G{r}")
    calc(ws4, r, 9, f"=H{r}*27.2")
    inp(ws4, r, 10, None)
    lbl(ws4, r, 11, "L leche/día o kg carne/año", bg_r)
    inp(ws4, r, 12, None)

r_tot_ent = 16
ws4.row_dimensions[r_tot_ent].height = 22
merge(ws4, r_tot_ent, 1, r_tot_ent, 7, "TOTAL FERMENTACIÓN ENTÉRICA", GR2, '000000', bold=True, sz=10)
calc(ws4, r_tot_ent, 8, f"=SUM(H6:H{r_tot_ent-1})")
calc(ws4, r_tot_ent, 9, f"=SUM(I6:I{r_tot_ent-1})")

# Sección B — Gestión de Estiércol
merge(ws4, 18, 1, 18, 14, "SECCIÓN B — GESTIÓN DE ESTIÉRCOL (CH4 + N2O)", S1_D, sz=10)
cols_est = [
    "Categoría Animal", "Sistema\nProductivo",
    "Cabezas\n[ref. Sec.A]\n[INGRESAR]",
    "Sistema de\ngestión\nestiércol",
    "% estiércol\npor sistema\n[INGRESAR]",
    "Temp.\npromedio\nanual (°C)\n[INGRESAR]",
    "FE CH4\nestiércol\n(kg CH4\n/cab/año)",
    "FE N2O\nestiércol\n(kg N2O\n/cab/año)",
    "CH4\n(kg/año)\n[CALC]",
    "N2O\n(kg/año)\n[CALC]",
    "CH4\n(kg CO2e\n/año)\n[CALC]",
    "N2O\n(kg CO2e\n/año)\n[CALC]",
    "TOTAL\ntCO2e/año\n[CALC]", "Observaciones"
]
for j, text in enumerate(cols_est):
    header(ws4, 19, j + 1, text, bg=S1_D, sz=8)

sistemas_est = [
    ("Vacas en ordeñe (tambo)", "Tambo", None, "Laguna anaeróbica", "100%", None, 16.0, 0.10),
    ("Novillos feedlot", "Feedlot", None, "Almac. sólido / corral", "100%", None, 1.0, 0.20),
    ("Novillitos feedlot", "Feedlot", None, "Almac. sólido / corral", "100%", None, 1.0, 0.20),
    ("Resto del rodeo (extensivo)", "Extensivo/pastoreo", None, "Pastoreo directo", "100%", None, 1.0, 0.07),
]
for i, (cat, sist, cabs, sistema, pct, temp, fe_ch4, fe_n2o) in enumerate(sistemas_est):
    r = 20 + i
    ws4.row_dimensions[r].height = 22
    bg_r = 'FFFFFF' if i % 2 == 0 else 'F9F9F9'
    lbl(ws4, r, 1, cat, bg_r, bold=True)
    lbl(ws4, r, 2, sist, bg_r)
    inp(ws4, r, 3, cabs)
    lbl(ws4, r, 4, sistema, bg_r)
    inp(ws4, r, 5, pct)
    inp(ws4, r, 6, temp)
    ref(ws4, r, 7, fe_ch4)
    ref(ws4, r, 8, fe_n2o)
    calc(ws4, r, 9, f"=C{r}*G{r}")
    calc(ws4, r, 10, f"=C{r}*H{r}")
    calc(ws4, r, 11, f"=I{r}*27.2")
    calc(ws4, r, 12, f"=J{r}*273")
    calc(ws4, r, 13, f"=(K{r}+L{r})/1000")
    inp(ws4, r, 14, None)

r_tot_est = 24
merge(ws4, r_tot_est, 1, r_tot_est, 8, "TOTAL GESTIÓN DE ESTIÉRCOL", GR2, '000000', bold=True, sz=10)
calc(ws4, r_tot_est, 9, f"=SUM(I20:I{r_tot_est-1})")
calc(ws4, r_tot_est, 10, f"=SUM(J20:J{r_tot_est-1})")
calc(ws4, r_tot_est, 11, f"=SUM(K20:K{r_tot_est-1})")
calc(ws4, r_tot_est, 12, f"=SUM(L20:L{r_tot_est-1})")
calc(ws4, r_tot_est, 13, f"=SUM(M20:M{r_tot_est-1})")

merge(ws4, 26, 1, 26, 14,
      "NOTA Fermentación: FE Tier 1 para clima templado (FAO/IPCC 2006). Mejorar a Tier 2 cuando se disponga de "
      "peso vivo promedio por categoría y calidad de dieta. "
      "NOTA Estiércol: FE por defecto IPCC 2006 Tabla 10-14. La laguna anaeróbica de tambo es la fuente "
      "de mayor incertidumbre — verificar temperatura operativa. "
      "Las cabezas de la Sección B deben coincidir con las de la Sección A.",
      'FFF2CC', '7F6000', bold=False, sz=9)

set_col_widths(ws4, [24, 16, 10, 22, 10, 10, 10, 10, 10, 10, 12, 12, 12, 18])

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 5: S1 — SUELOS AGRÍCOLAS
# ─────────────────────────────────────────────────────────────────────────────
ws5 = wb.create_sheet("S1-4 SUELOS AGRICOLAS")
ws5.sheet_properties.tabColor = S1_D
ws5.freeze_panes = 'A5'

merge(ws5, 1, 1, 1, 14, "SCOPE 1 — SUELOS AGRÍCOLAS  |  N2O de Fertilizantes y Urea", S1_D, sz=13)
merge(ws5, 2, 1, 2, 14,
      "IPCC 2006 Vol.4 Cap.11  |  EF1 (N directo) = 0,01 kg N2O-N/kg N aplicado  |  "
      "Emisión CO2 urea: 0,734 kg CO2/kg urea  |  EF3 (nitrificación/desnitrif.) incluido en cálculo indirecto",
      S1_L, '000000', bold=False, sz=9)

cols_suel = [
    "Cultivo / Lote", "Campaña",
    "Superficie\n(ha)\n[INGRESAR]",
    "Fertil. N\naplicado\n(kg N/ha)\n[INGRESAR]",
    "N total\naplicado\n(kg N)\n[CALC]",
    "Urea\naplicada\n(kg/ha)\n[INGRESAR]",
    "Urea total\n(kg)\n[CALC]",
    "Residuos\ncultivo\nincorpor.\n(t/ha)\n[INGRESAR]",
    "N en\nresiduos\n(kg N)\n[CALC]",
    "EF1\nN2O\ndirecto",
    "N2O\ndirecto\n(kg/año)\n[CALC]",
    "CO2 hidr.\nurea\n(kg CO2)\n[CALC]",
    "N2O total\n(kg CO2e/\naño)\n[CALC]",
    "TOTAL\ntCO2e/año\n[CALC]"
]
for j, text in enumerate(cols_suel):
    header(ws5, 4, j + 1, text, bg=S1_D, sz=8)

cultivos = [
    ("Soja (1ra)", ""),
    ("Soja (2da)", ""),
    ("Maíz", ""),
    ("Trigo", ""),
    ("Sorgo forrajero", ""),
    ("Girasol", ""),
    ("Pasturas implantadas", ""),
    ("Otro cultivo 1", ""),
    ("Otro cultivo 2", ""),
]
for i, (cult, camp) in enumerate(cultivos):
    r = 5 + i
    ws5.row_dimensions[r].height = 20
    bg_r = 'FFFFFF' if i % 2 == 0 else 'F9F9F9'
    lbl(ws5, r, 1, cult, bg_r, bold=True)
    inp(ws5, r, 2, camp)
    inp(ws5, r, 3, None)
    inp(ws5, r, 4, None)
    calc(ws5, r, 5, f"=C{r}*D{r}")
    inp(ws5, r, 6, None)
    calc(ws5, r, 7, f"=C{r}*F{r}")
    inp(ws5, r, 8, None)
    calc(ws5, r, 9, f"=C{r}*H{r}*0.006")   # 0.6% N en residuos (IPCC default)
    ref(ws5, r, 10, 0.01)                   # EF1
    # N2O directo = (N_fert + N_residuos) * EF1 * 44/28
    calc(ws5, r, 11, f"=(E{r}+I{r})*J{r}*(44/28)")
    # CO2 hidrolisis urea = 0.734 kg CO2/kg urea
    calc(ws5, r, 12, f"=G{r}*0.734")
    # N2O total CO2e
    calc(ws5, r, 13, f"=K{r}*273")
    # Total tCO2e (N2O + CO2 urea)
    calc(ws5, r, 14, f"=(L{r}+M{r})/1000")

r_tot_suel = 14
ws5.row_dimensions[r_tot_suel].height = 22
merge(ws5, r_tot_suel, 1, r_tot_suel, 9, "TOTAL SUELOS AGRÍCOLAS", GR2, '000000', bold=True, sz=10)
calc(ws5, r_tot_suel, 10, "---")
calc(ws5, r_tot_suel, 11, f"=SUM(K5:K{r_tot_suel-1})")
calc(ws5, r_tot_suel, 12, f"=SUM(L5:L{r_tot_suel-1})")
calc(ws5, r_tot_suel, 13, f"=SUM(M5:M{r_tot_suel-1})")
calc(ws5, r_tot_suel, 14, f"=SUM(N5:N{r_tot_suel-1})")

# Agroquímicos y otros insumos
merge(ws5, 16, 1, 16, 14, "SECCIÓN B — EMISIONES ADICIONALES: AGROQUÍMICOS Y ENMIENDAS", S1_D, sz=10)
merge(ws5, 17, 1, 17, 14,
      "Los agroquímicos (herbicidas, fungicidas, insecticidas) no tienen FE estandarizado en IPCC para Scope 1. "
      "Se registran aquí para Scope 3 (emisiones upstream de fabricación) cuando se reporte esa categoría. "
      "Enmiendas calcáreas (cal, dolomita): FE = 0,44 kg CO2/kg cal; 0,477 kg CO2/kg dolomita.",
      'FFF2CC', '7F6000', bold=False, sz=9)

cols_insumos = [
    "Insumo", "Tipo", "Cantidad\nAnual (kg o L)\n[INGRESAR]", "Unidad",
    "FE referencial\n(kg CO2/kg o L)", "Emisiones CO2\n(kg/año)\n[CALC]",
    "Scope", "Observaciones"
]
for j, text in enumerate(cols_insumos):
    header(ws5, 18, j + 1, text, bg=GR3, fg='000000', sz=9)

insumos = [
    ("Urea", "Fertilizante N", None, "kg", "Ya incluido en S A", "---", "Scope 1", "Hidrolisis: 0,734 kg CO2/kg"),
    ("Fertilizante N (otro)", "Fertilizante N", None, "kg N", "---", "---", "Scope 1", "N directo en Sec A"),
    ("Herbicidas", "Agroquímico", None, "L", "No estandarizado", "---", "Scope 3", "Registrar para futuro Sc3"),
    ("Fungicidas", "Agroquímico", None, "L", "No estandarizado", "---", "Scope 3", "Registrar para futuro Sc3"),
    ("Insecticidas", "Agroquímico", None, "L", "No estandarizado", "---", "Scope 3", "Registrar para futuro Sc3"),
    ("Cal agrícola", "Enmienda", None, "kg", "0,440", None, "Scope 1", "Encalado: 0,44 kg CO2/kg"),
    ("Dolomita", "Enmienda", None, "kg", "0,477", None, "Scope 1", "Si se usa dolomita"),
]
for i, (ins, tipo, cant, unidad, fe, ems, scope, obs) in enumerate(insumos):
    r = 19 + i
    ws5.row_dimensions[r].height = 20
    bg_r = 'FFFFFF' if i % 2 == 0 else 'F9F9F9'
    lbl(ws5, r, 1, ins, bg_r, bold=True)
    lbl(ws5, r, 2, tipo, bg_r)
    inp(ws5, r, 3, cant)
    lbl(ws5, r, 4, unidad, bg_r)
    ref(ws5, r, 5, fe)
    if ems == "---":
        set_cell(ws5, r, 6, "---", bg='E2EFDA', fg='375623', halign='center', italic=True)
    elif ems is None:
        calc(ws5, r, 6, f"=C{r}*E{r}" if fe not in ("No estandarizado", "---") else "---")
    lbl(ws5, r, 7, scope, bg_r)
    lbl(ws5, r, 8, obs, bg_r)

set_col_widths(ws5, [22, 16, 11, 10, 12, 12, 10, 14, 10, 8, 12, 14, 14, 12])

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 6: S1 — EFLUENTES
# ─────────────────────────────────────────────────────────────────────────────
ws6 = wb.create_sheet("S1-5 EFLUENTES")
ws6.sheet_properties.tabColor = S1_D
ws6.freeze_panes = 'A5'

merge(ws6, 1, 1, 1, 12, "SCOPE 1 — EFLUENTES LÍQUIDOS  |  Feedlot + Tambo", S1_D, sz=13)
merge(ws6, 2, 1, 2, 12,
      "IPCC 2006 Vol.5 Cap.6  |  Bo (CH4 máx.) feedlot: 0,16 m³ CH4/kg DBO  |  "
      "EF efluentes N2O (Doering): 0,005 kg N2O-N/kg N efluente  |  Densidad CH4: 0,00067 t/m³",
      S1_L, '000000', bold=False, sz=9)

cols_efl = [
    "Origen\nEfluente", "Sistema de\nTratamiento",
    "Caudal\ndiario\n(m³/día)\n[INGRESAR]",
    "Días\noperación\n/año\n[INGRESAR]",
    "Volumen\nanual\n(m³/año)\n[CALC]",
    "DBO\nafluente\n(mg/L)\n[INGRESAR]",
    "DBO\nEfluente\n(mg O2/L)\n[INGRESAR]",
    "Bo\n(m³ CH4\n/kg DBO)",
    "MCF\n(Factor\ncorr.)",
    "CH4\n(m³/año)\n[CALC]",
    "CH4\n(kg CO2e\n/año)\n[CALC]",
    "TOTAL\ntCO2e/año\n[CALC]"
]
for j, text in enumerate(cols_efl):
    header(ws6, 4, j + 1, text, bg=S1_D, sz=8)

efluentes = [
    ("Tambo — efluente sala ordeñe", "Laguna anaeróbica", None, "365", None, None, None, 0.16, 0.8),
    ("Tambo — corrales y aguadas", "Laguna/infiltración", None, "365", None, None, None, 0.16, 0.3),
    ("Feedlot — escorrentía corrales", "Laguna anaeróbica", None, "365", None, None, None, 0.16, 0.8),
    ("Feedlot — aguas de bebederos", "Infiltración", None, "200", None, None, None, 0.08, 0.1),
]
for i, (orig, sist, caudal, dias, vol, dbo_in, dbo_out, bo, mcf) in enumerate(efluentes):
    r = 5 + i
    ws6.row_dimensions[r].height = 22
    bg_r = 'FFFFFF' if i % 2 == 0 else 'F9F9F9'
    lbl(ws6, r, 1, orig, bg_r, bold=True)
    lbl(ws6, r, 2, sist, bg_r)
    inp(ws6, r, 3, caudal)
    inp(ws6, r, 4, dias)
    calc(ws6, r, 5, f"=C{r}*D{r}")
    inp(ws6, r, 6, dbo_in)
    inp(ws6, r, 7, dbo_out)
    ref(ws6, r, 8, bo)
    ref(ws6, r, 9, mcf)
    # CH4 = Vol(m³) * (DBO_in - DBO_out) g/m³ /1000 kg/g * Bo * MCF
    calc(ws6, r, 10, f"=E{r}*(F{r}-G{r})/1000*H{r}*I{r}")
    calc(ws6, r, 11, f"=J{r}*0.00067*1000*27.2")
    calc(ws6, r, 12, f"=K{r}/1000")

r_tot_efl = 9
ws6.row_dimensions[r_tot_efl].height = 22
merge(ws6, r_tot_efl, 1, r_tot_efl, 9, "TOTAL EFLUENTES", GR2, '000000', bold=True, sz=10)
calc(ws6, r_tot_efl, 10, f"=SUM(J5:J{r_tot_efl-1})")
calc(ws6, r_tot_efl, 11, f"=SUM(K5:K{r_tot_efl-1})")
calc(ws6, r_tot_efl, 12, f"=SUM(L5:L{r_tot_efl-1})")

# Datos de apoyo para tambo
merge(ws6, 11, 1, 11, 12, "DATOS DE APOYO — TAMBO (completar para contexto)", S1_D, sz=10)
tambo_fields = [
    ("Vacas en ordeñe (promedio año)", None, "cabezas"),
    ("Producción leche por vaca (L/día)", None, "L/vaca/día"),
    ("Producción leche total (L/año)", None, "L/año — o CALC si los anteriores están completos"),
    ("Días de ordeñe por año", None, "días"),
    ("Tipo de sala de ordeñe", None, "rotativa, espina de pescado, etc."),
    ("Sistema de lavado de sala", None, "agua fría/caliente, CIP, etc."),
    ("Destino del efluente de sala", None, "laguna, compost, biodigestor, etc."),
]
for i, (field, val, obs) in enumerate(tambo_fields):
    r = 12 + i
    ws6.row_dimensions[r].height = 20
    lbl(ws6, r, 1, field, GR1, bold=True)
    ws6.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    inp(ws6, r, 2, val)
    ws6.merge_cells(start_row=r, start_column=5, end_row=r, end_column=12)
    lbl(ws6, r, 5, obs, GR1)

# Datos de apoyo para feedlot
merge(ws6, 20, 1, 20, 12, "DATOS DE APOYO — FEEDLOT (completar para contexto)", S1_D, sz=10)
feedlot_fields = [
    ("Cabezas en engorde simultáneamente (promedio)", None, "cabezas"),
    ("Días promedio de estadía en feedlot", None, "días"),
    ("Entradas de hacienda por año (ciclos)", None, "ciclos/año"),
    ("Superficie total de corrales (m²)", None, "m²"),
    ("Sistema de drenaje", None, "describirlo"),
    ("Destino del estiércol sólido", None, "compost, esparcido, venta, etc."),
    ("Capacidad laguna de efluentes (m³)", None, "m³"),
]
for i, (field, val, obs) in enumerate(feedlot_fields):
    r = 21 + i
    ws6.row_dimensions[r].height = 20
    lbl(ws6, r, 1, field, GR1, bold=True)
    ws6.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    inp(ws6, r, 2, val)
    ws6.merge_cells(start_row=r, start_column=5, end_row=r, end_column=12)
    lbl(ws6, r, 5, obs, GR1)

set_col_widths(ws6, [28, 22, 10, 9, 11, 10, 10, 9, 8, 12, 14, 12])

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 7: S2 — ELECTRICIDAD
# ─────────────────────────────────────────────────────────────────────────────
ws7 = wb.create_sheet("S2-1 ELECTRICIDAD")
ws7.sheet_properties.tabColor = S2_D
ws7.freeze_panes = 'A5'

merge(ws7, 1, 1, 1, 10, "SCOPE 2 — ELECTRICIDAD DE RED  |  Enfoque Basado en Localización", S2_D, sz=13)
merge(ws7, 2, 1, 2, 10,
      "GHG Protocol — Scope 2 Guidance (2015)  |  Factor de Emisión Red Argentina CAMMESA 2023: 0,383 kg CO2e/kWh "
      "(⚠ verificar valor vigente del año del inventario en cammesaweb.cammesa.com)  |  "
      "20% del consumo eléctrico total es de red; 80% autogenerado (incluido en S1-2)",
      S2_L, '1F3864', bold=False, sz=9)

cols_elec = [
    "Mes", "Año",
    "kWh\nfacturados\n(red)\n[INGRESAR]",
    "Demanda\npunta\n(kW)\n[INGRESAR]",
    "Costo\n($)\n[INGRESAR]",
    "N° Factura\n[INGRESAR]",
    "FE Red\n(kg CO2e\n/kWh)",
    "Emisiones\nCO2e\n(kg)\n[CALC]",
    "Emisiones\nCO2e\n(tCO2e)\n[CALC]",
    "Observaciones"
]
for j, text in enumerate(cols_elec):
    header(ws7, 4, j + 1, text, bg=S2_D, sz=9)

meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
         "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
for i, mes in enumerate(meses):
    r = 5 + i
    ws7.row_dimensions[r].height = 20
    bg_r = 'FFFFFF' if i % 2 == 0 else 'F9F9F9'
    lbl(ws7, r, 1, mes, bg_r)
    inp(ws7, r, 2, None)   # Año
    inp(ws7, r, 3, None)   # kWh
    inp(ws7, r, 4, None)   # Demanda punta
    inp(ws7, r, 5, None)   # Costo
    inp(ws7, r, 6, None)   # N° Factura
    ref(ws7, r, 7, 0.383)
    calc(ws7, r, 8, f"=C{r}*G{r}")
    calc(ws7, r, 9, f"=H{r}/1000")
    inp(ws7, r, 10, None)  # Obs

r_tot_elec = 17
ws7.row_dimensions[r_tot_elec].height = 22
merge(ws7, r_tot_elec, 1, r_tot_elec, 2, "TOTAL ANUAL", S2_L, '1F3864', bold=True, sz=10)
calc(ws7, r_tot_elec, 3, "=SUM(C5:C16)")
merge(ws7, r_tot_elec, 4, r_tot_elec, 6, "ELECTRICIDAD DE RED", S2_L, '1F3864', bold=True, sz=10)
calc(ws7, r_tot_elec, 8, "=SUM(H5:H16)")
calc(ws7, r_tot_elec, 9, "=SUM(I5:I16)")

# Autogeneración (informativo)
merge(ws7, 19, 1, 19, 10, "SECCIÓN B — AUTOGENERACIÓN (referencia cruzada con S1-2)", S2_D, sz=10)
auto_fields = [
    ("kWh generados por generadores (año)", None, "De hoja S1-2 col H"),
    ("kWh generados por otra fuente renovable", None, "Solar, eólico, etc."),
    ("Total kWh consumidos en establecimiento (red + autogen.)", None, "Suma para contexto energético"),
    ("% autogeneración sobre consumo total", None, "Calc: autogen / (red + autogen)"),
]
for j, text in enumerate(["Descripción", "Valor", "Fuente / Nota"]):
    header(ws7, 20, j + 1, text, bg=S2_D, sz=9)

for i, (desc, val, note) in enumerate(auto_fields):
    r = 21 + i
    ws7.row_dimensions[r].height = 20
    lbl(ws7, r, 1, desc, GR1, bold=True)
    ws7.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    inp(ws7, r, 2, val)
    ws7.merge_cells(start_row=r, start_column=4, end_row=r, end_column=10)
    lbl(ws7, r, 4, note, GR1)

merge(ws7, 26, 1, 26, 10,
      "NOTA: Adjuntar copia de las 12 facturas eléctricas al expediente del inventario (requisito ISO 14064). "
      "Si no se dispone de medidor separado para carga de red, solicitar a la distribuidora histórico de consumo. "
      "Usar enfoque de market-based si se tiene contrato de energía renovable certificada (GO/REC).",
      'DCE6F1', '1F3864', bold=False, sz=9)

set_col_widths(ws7, [12, 8, 14, 12, 14, 16, 10, 14, 12, 22])

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 8: SUMIDEROS
# ─────────────────────────────────────────────────────────────────────────────
ws8 = wb.create_sheet("SUMIDEROS")
ws8.sheet_properties.tabColor = SK_D
ws8.freeze_panes = 'A5'

merge(ws8, 1, 1, 1, 12, "SUMIDEROS DE CARBONO  |  Monte Nativo + Pasturas", SK_D, sz=13)
merge(ws8, 2, 1, 2, 12,
      "IPCC 2006 Vol.4 Cap.4/6  |  Valores negativos = capturas (CO2 removido de la atmósfera)  |  "
      "Tier 1: factores por defecto por bioma/región. Para Tier 2 se requiere inventario forestal.",
      SK_L, '000000', bold=False, sz=9)

# Sección A — Monte Nativo / Bosque
merge(ws8, 3, 1, 3, 12, "SECCIÓN A — MONTE NATIVO / VEGETACIÓN PERENNE", SK_D, sz=10)
cols_mont = [
    "Unidad de\nVegetación", "Tipo de\nFormación",
    "Superficie\n(ha)\n[INGRESAR]",
    "Estado de\nConservación\n[INGRESAR]",
    "Categoría\nOrdenamiento\nTerritorial\n(Ley 26.331)",
    "Biomasa\naérea\nactual\n(t MS/ha)\n[INGRESAR]",
    "Factor de\nCaptura\nTier 1\n(tCO2e/ha/año)",
    "Captura\nneta\n(tCO2e/año)\n[CALC]",
    "¿Hay\ndeforest.?\n[INGRESAR]",
    "ha\ndeforest.\neste año\n[INGRESAR]",
    "Emisión\ndeforest.\n(tCO2e)\n[CALC]",
    "Neto\n(tCO2e/año)\n[CALC]"
]
for j, text in enumerate(cols_mont):
    header(ws8, 4, j + 1, text, bg=SK_D, sz=8)

monte_units = [
    ("Bosque nativo sector Norte", "Bosque chaqueño / serrano", None, None, "Rojo / Amarillo", None, 3.2),
    ("Bosque nativo sector Sur", "Bosque chaqueño / serrano", None, None, "Rojo / Amarillo", None, 3.2),
    ("Monte ribereño / galería", "Bosque ribereño", None, None, "Rojo", None, 4.5),
    ("Arbustales / malezales", "Matorral xerófilo", None, None, "Amarillo", None, 1.5),
]
for i, (nombre, tipo, sup, estado, cat, bio, fe) in enumerate(monte_units):
    r = 5 + i
    ws8.row_dimensions[r].height = 22
    bg_r = 'FFFFFF' if i % 2 == 0 else 'F9F9F9'
    lbl(ws8, r, 1, nombre, bg_r, bold=True)
    lbl(ws8, r, 2, tipo, bg_r)
    inp(ws8, r, 3, sup)
    inp(ws8, r, 4, estado)
    lbl(ws8, r, 5, cat, bg_r)
    inp(ws8, r, 6, bio)
    ref(ws8, r, 7, fe)
    calc(ws8, r, 8, f"=-C{r}*G{r}")
    inp(ws8, r, 9, "No")
    inp(ws8, r, 10, None)
    calc(ws8, r, 11, f"=J{r}*G{r}*3.67")
    calc(ws8, r, 12, f"=H{r}+K{r}")

r_tot_mont = 9
merge(ws8, r_tot_mont, 1, r_tot_mont, 6, "SUBTOTAL MONTE NATIVO", GR2, '000000', bold=True, sz=10)
calc(ws8, r_tot_mont, 8, f"=SUM(H5:H{r_tot_mont-1})")
calc(ws8, r_tot_mont, 11, f"=SUM(K5:K{r_tot_mont-1})")
calc(ws8, r_tot_mont, 12, f"=SUM(L5:L{r_tot_mont-1})")

# Sección B — Pasturas
merge(ws8, 11, 1, 11, 12, "SECCIÓN B — PASTURAS IMPLANTADAS Y SUELO BAJO AGRICULTURA", SK_D, sz=10)
cols_past = [
    "Tipo de Pastura /\nCultivo de Cobertura", "Manejo",
    "Superficie\n(ha)\n[INGRESAR]",
    "Años de\nestablecida\n[INGRESAR]",
    "Estado\nMasomenos\n[INGRESAR]",
    "Labranza\npredominante\n[INGRESAR]",
    "FE Stock C\nsuelo\n(tCO2e/ha/año)\n[Tier1]",
    "Captura\nneta\n(tCO2e/año)\n[CALC]",
    "Observaciones", "", "", ""
]
for j, text in enumerate(cols_past):
    header(ws8, 12, j + 1, text, bg=SK_D, sz=8)

pasturas = [
    ("Pasturas implantadas perennes", "Pastoreo rotativo", None, None, None, "Sin labranza", 0.5),
    ("Pasturas anuales / verdeos", "Pastoreo continuo", None, None, None, "Siembra directa", 0.2),
    ("Cultivos anuales (soja/maíz/trigo)", "Agricultura", None, None, None, "Siembra directa", 0.3),
    ("Rastrojos incorporados", "Cobertura", None, None, None, "SD / Mín. labranza", 0.15),
]
for i, (tipo, manejo, sup, anios, estado, labor, fe) in enumerate(pasturas):
    r = 13 + i
    ws8.row_dimensions[r].height = 22
    bg_r = 'FFFFFF' if i % 2 == 0 else 'F9F9F9'
    lbl(ws8, r, 1, tipo, bg_r, bold=True)
    lbl(ws8, r, 2, manejo, bg_r)
    inp(ws8, r, 3, sup)
    inp(ws8, r, 4, anios)
    inp(ws8, r, 5, estado)
    inp(ws8, r, 6, labor)
    ref(ws8, r, 7, fe)
    calc(ws8, r, 8, f"=-C{r}*G{r}")
    ws8.merge_cells(start_row=r, start_column=9, end_row=r, end_column=12)
    inp(ws8, r, 9, None)

r_tot_past = 17
merge(ws8, r_tot_past, 1, r_tot_past, 6, "SUBTOTAL PASTURAS Y SUELOS", GR2, '000000', bold=True, sz=10)
calc(ws8, r_tot_past, 8, f"=SUM(H13:H{r_tot_past-1})")

r_tot_sk = 19
merge(ws8, r_tot_sk, 1, r_tot_sk, 6, "TOTAL SUMIDEROS (Monte + Pasturas)", SK_L, '000000', bold=True, sz=11)
calc(ws8, r_tot_sk, 8, f"=H{r_tot_mont}+H{r_tot_past}")
calc(ws8, r_tot_sk, 12, f"=L{r_tot_mont}")

merge(ws8, 21, 1, 21, 12,
      "NOTA: Valores NEGATIVOS = capturas netas (el establecimiento absorbe CO2). Valores POSITIVOS = fuentes. "
      "Para inventario forestal verificado (VCS/Carbono Neutro) se requiere medición in situ (Tier 2/3). "
      "Factor de deforestación: 3,67 = relación masa molar CO2/C. "
      "Ley 26.331: bosques Rojo = no modificar; Amarillo = uso sostenible; Verde = aprovechamiento permitido. "
      "Los pastizales de siembra directa con más de 5 años tienden a ser sumideros netos.",
      'C6EFCE', '375623', bold=False, sz=9)

set_col_widths(ws8, [26, 20, 10, 10, 14, 16, 12, 14, 14, 10, 14, 14])

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 9: FACTORES DE EMISIÓN
# ─────────────────────────────────────────────────────────────────────────────
ws9 = wb.create_sheet("FACTORES EMISION")
ws9.sheet_properties.tabColor = GR3

merge(ws9, 1, 1, 1, 8, "TABLA DE FACTORES DE EMISIÓN DE REFERENCIA", NAVY, sz=13)
merge(ws9, 2, 1, 2, 8,
      "Fuentes: IPCC 2006 | GHG Protocol | CAMMESA 2023 | AR6 IPCC 2021 — NO modificar esta hoja. Solo lectura.",
      GR2, '000000', bold=False, sz=9)

secciones_fe = [
    ("COMBUSTIBLES", [
        ("Gasoil / Diésel", "Combustión móvil y estacionaria", "L", "2,701", "0,100", "0,600", "IPCC 2006 Tabla 1.4"),
        ("GLP / Propano (Zeppelin)", "Combustión estacionaria", "kg", "2,983", "0,010", "0,100", "IPCC 2006 Tabla 1.4"),
        ("Propano líquido", "Combustión estacionaria", "L", "1,534", "0,010", "0,100", "IPCC 2006 Tabla 1.4"),
        ("Gasolina / Nafta", "Combustión móvil", "L", "2,296", "0,100", "0,600", "IPCC 2006 Tabla 1.4"),
        ("Gas natural", "Combustión estacionaria", "m³", "1,879", "0,005", "0,010", "IPCC 2006 Tabla 1.4"),
    ]),
    ("ELECTRICIDAD", [
        ("Red eléctrica Argentina (CAMMESA 2023)", "Scope 2 — factor de emisión localización", "kWh", "0,383", "—", "—", "CAMMESA — Factor de Emisión 2023"),
        ("Red eléctrica Argentina (CAMMESA 2022)", "Referencia histórica", "kWh", "0,404", "—", "—", "CAMMESA — Factor de Emisión 2022"),
        ("⚠ Verificar vigencia anual", "Actualizar con el valor publicado del año del inventario", "kWh", "—", "—", "—", "cammesaweb.cammesa.com/download/factor-de-emision"),
    ]),
    ("GANADERÍA — FERMENTACIÓN ENTÉRICA (Tier 1, clima templado)", [
        ("Vacas lecheras (>200 L/día producción)", "Fermentación entérica", "cab/año", "119,8 kg CH4", "—", "—", "IPCC 2006 Tabla 10-11"),
        ("Bovinos adultos carne (>300 kg)", "Fermentación entérica", "cab/año", "56,5 kg CH4", "—", "—", "IPCC 2006 Tabla 10-11"),
        ("Novillo / vaquillona recría (200-300 kg)", "Fermentación entérica", "cab/año", "47,8 kg CH4", "—", "—", "IPCC 2006 Tabla 10-11"),
        ("Terneros < 1 año", "Fermentación entérica", "cab/año", "27,0 kg CH4", "—", "—", "IPCC 2006 Tabla 10-11"),
        ("Toros reproductores", "Fermentación entérica", "cab/año", "73,0 kg CH4", "—", "—", "IPCC 2006 Tabla 10-11"),
        ("Ovinos (adultos)", "Fermentación entérica", "cab/año", "8,0 kg CH4", "—", "—", "IPCC 2006 Tabla 10-11"),
    ]),
    ("GANADERÍA — GESTIÓN DE ESTIÉRCOL (Tier 1)", [
        ("Vacas lecheras — laguna anaeróbica", "CH4 estiércol", "cab/año", "16,0 kg CH4", "—", "0,10 kg N2O", "IPCC 2006 Tabla 10-14/15"),
        ("Bovinos carne — almac. sólido/corral", "CH4 estiércol", "cab/año", "1,0 kg CH4", "—", "0,20 kg N2O", "IPCC 2006 Tabla 10-14/15"),
        ("Bovinos — pastoreo directo", "CH4 estiércol", "cab/año", "1,0 kg CH4", "—", "0,07 kg N2O", "IPCC 2006 Tabla 10-14/15"),
    ]),
    ("SUELOS AGRÍCOLAS", [
        ("Fertilizante N (nitrogenado)", "N2O directo suelos EF1", "kg N aplic.", "0,01 kg N2O-N", "—", "—", "IPCC 2006 Tabla 11-1"),
        ("Urea aplicada", "CO2 hidrólisis", "kg urea", "0,734 kg CO2", "—", "—", "IPCC 2006 Vol.4 Cap.11"),
        ("Residuos de cultivos incorporados", "N2O indirecto", "kg N residuos", "0,01 kg N2O-N", "—", "—", "IPCC 2006 Tabla 11-1"),
    ]),
    ("POTENCIALES DE CALENTAMIENTO GLOBAL (AR6 — 100 años)", [
        ("CH4 fósil (combustión gasoil/GLP)", "GWP base CO2 — Hojas S1-1 y S1-2", "kg CH4", "29,8 kg CO2e", "—", "—", "IPCC AR6 2021 Tabla 7.SM.7"),
        ("CH4 biogénico (ganadería/efluentes)", "GWP base CO2 — Hojas S1-3 y S1-5", "kg CH4", "27,2 kg CO2e", "—", "—", "IPCC AR6 2021 Tabla 7.SM.7"),
        ("N2O (óxido nitroso)", "GWP base CO2", "kg N2O", "273 kg CO2e", "—", "—", "IPCC AR6 2021 Tabla 7.SM.7"),
        ("CO2 (dióxido de carbono)", "GWP base CO2", "kg CO2", "1 kg CO2e", "—", "—", "Línea base"),
    ]),
]

r_fe = 4
for seccion, filas in secciones_fe:
    ws9.row_dimensions[r_fe].height = 22
    merge(ws9, r_fe, 1, r_fe, 8, seccion, GR3, '000000', bold=True, sz=10)
    r_fe += 1
    for j, text in enumerate(["Gas / Fuente", "Aplicación", "Unidad", "FE CO2 (kg CO2/unidad)", "FE CH4 (g/unidad)", "FE N2O (g/unidad)", "Fuente"]):
        header(ws9, r_fe, j + 1, text, bg=NAVY, sz=9)
    ws9.merge_cells(start_row=r_fe, start_column=7, end_row=r_fe, end_column=8)
    r_fe += 1
    for i, fila in enumerate(filas):
        ws9.row_dimensions[r_fe].height = 18
        bg_r = 'FFFFFF' if i % 2 == 0 else 'F9F9F9'
        for j_c, val in enumerate(fila):
            if j_c == 0:
                set_cell(ws9, r_fe, j_c + 1, val, bg=bg_r, bold=True, halign='left')
            elif j_c == 6:
                ws9.merge_cells(start_row=r_fe, start_column=7, end_row=r_fe, end_column=8)
                set_cell(ws9, r_fe, 7, val, bg='BDD7EE', halign='center', sz=8, italic=True)
            elif j_c in (3, 4, 5):
                ref(ws9, r_fe, j_c + 1, val)
            else:
                set_cell(ws9, r_fe, j_c + 1, val, bg=bg_r, halign='center')
        r_fe += 1
    r_fe += 1

set_col_widths(ws9, [36, 28, 14, 18, 14, 14, 20, 20])

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 10: RESUMEN / DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────
ws10 = wb.create_sheet("RESUMEN DASHBOARD")
ws10.sheet_properties.tabColor = OR_D
ws10.freeze_panes = 'A4'

merge(ws10, 1, 1, 1, 10,
      "RESUMEN DEL INVENTARIO DE GASES DE EFECTO INVERNADERO  |  HUELLA DE CARBONO CORPORATIVA",
      NAVY, sz=14)
merge(ws10, 2, 1, 2, 10,
      "Este resumen se actualiza automáticamente al completar las hojas de datos. "
      "Usar esta hoja como fuente del dashboard interactivo.",
      GR2, '000000', bold=False, sz=9)

# Tabla resumen por fuente
merge(ws10, 4, 1, 4, 10, "TABLA 1 — EMISIONES Y CAPTURAS POR CATEGORÍA", NAVY, sz=11)
cols_res = [
    "Código", "Categoría", "Scope",
    "CO2\n(tCO2/año)", "CH4\n(tCO2e/año)", "N2O\n(tCO2e/año)",
    "TOTAL\ntCO2e/año",
    "% del\ntotal\nemisiones",
    "Hoja de\nReferencia",
    "Observaciones"
]
for j, text in enumerate(cols_res):
    header(ws10, 5, j + 1, text, bg=NAVY, sz=9)

categorias_res = [
    ("S1-1", "Combustión Móvil (maquinaria)", "Scope 1", None, None, None, S1_L),
    ("S1-2", "Combustión Estacionaria (gen. + zeppelin)", "Scope 1", None, None, None, S1_L),
    ("S1-3", "Fermentación Entérica", "Scope 1", None, None, None, S1_L),
    ("S1-4", "Gestión de Estiércol", "Scope 1", None, None, None, S1_L),
    ("S1-5", "Suelos Agrícolas (N2O + CO2 urea)", "Scope 1", None, None, None, S1_L),
    ("S1-6", "Efluentes (feedlot + tambo)", "Scope 1", None, None, None, S1_L),
    ("S2-1", "Electricidad de Red", "Scope 2", None, None, None, S2_L),
]
refs_sheets = [
    "'S1-1 MAQUINARIA'!O14",
    "'S1-2 ENERGIA ESTAC'!L14",
    "'S1-3 GANADERIA'!I16",
    "'S1-3 GANADERIA'!M24",
    "'S1-4 SUELOS AGRICOLAS'!N14",
    "'S1-5 EFLUENTES'!L9",
    "'S2-1 ELECTRICIDAD'!I17",
]
for i, ((code, name, scope, co2, ch4, n2o, bg), ref_sheet) in enumerate(zip(categorias_res, refs_sheets)):
    r = 6 + i
    ws10.row_dimensions[r].height = 22
    set_cell(ws10, r, 1, code, bg=bg, bold=True, halign='center')
    lbl(ws10, r, 2, name, bg)
    set_cell(ws10, r, 3, scope, bg=bg, halign='center')
    # For now, reference the total tCO2e column from each sheet
    calc(ws10, r, 4, f"=IFERROR({ref_sheet},0)")
    set_cell(ws10, r, 5, "ver hoja", bg='E2EFDA', fg='375623', italic=True, halign='center', sz=9)
    set_cell(ws10, r, 6, "ver hoja", bg='E2EFDA', fg='375623', italic=True, halign='center', sz=9)
    calc(ws10, r, 7, f"=D{r}")
    # % will be calculated after total row
    set_cell(ws10, r, 8, f"=IFERROR(G{r}/G13*100,0)", bg='E2EFDA', fg='375623', italic=True, halign='center')
    set_cell(ws10, r, 9, ref_sheet.split('!')[0].replace("'", ""), bg=GR1, halign='center', sz=9)
    inp(ws10, r, 10, None)

# Subtotal emisiones
r_sub_em = 13
ws10.row_dimensions[r_sub_em].height = 24
merge(ws10, r_sub_em, 1, r_sub_em, 3, "SUBTOTAL EMISIONES (S1 + S2)", S1_L, '000000', bold=True, sz=10)
calc(ws10, r_sub_em, 7, f"=SUM(G6:G12)")

# Sumideros
merge(ws10, 15, 1, 15, 10, "TABLA 2 — CAPTURAS (SUMIDEROS)", SK_D, sz=11)
for j, text in enumerate(cols_res):
    header(ws10, 16, j + 1, text, bg=SK_D, sz=9)

sumideros_res = [
    ("SK-1", "Monte Nativo", "Sumidero", "SUMIDEROS!H9"),
    ("SK-2", "Pasturas Implantadas", "Sumidero", "SUMIDEROS!H17"),
]
for i, (code, name, tipo, ref_sk) in enumerate(sumideros_res):
    r = 17 + i
    ws10.row_dimensions[r].height = 20
    set_cell(ws10, r, 1, code, bg=SK_L, bold=True, halign='center')
    lbl(ws10, r, 2, name, SK_L)
    set_cell(ws10, r, 3, tipo, bg=SK_L, halign='center')
    ws10.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
    calc(ws10, r, 4, f"=IFERROR({ref_sk},0)")
    calc(ws10, r, 7, f"=D{r}")
    set_cell(ws10, r, 9, ref_sk.split('!')[0], bg=GR1, halign='center', sz=9)
    inp(ws10, r, 10, None)

r_sub_sk = 19
merge(ws10, r_sub_sk, 1, r_sub_sk, 3, "SUBTOTAL CAPTURAS (Sumideros)", SK_L, '000000', bold=True, sz=10)
calc(ws10, r_sub_sk, 7, "=SUM(G17:G18)")

# Huella neta
r_net = 21
ws10.row_dimensions[r_net].height = 30
merge(ws10, r_net, 1, r_net, 3, "HUELLA DE CARBONO NETA (Emisiones + Capturas)", NAVY, sz=12)
calc(ws10, r_net, 7, f"=G{r_sub_em}+G{r_sub_sk}")
set_cell(ws10, r_net, 8, "tCO2e/año", bg=NAVY, fg='FFFFFF', bold=True, halign='center')

# KPIs para dashboard
merge(ws10, 23, 1, 23, 10, "TABLA 3 — INDICADORES DE INTENSIDAD (KPIs para Dashboard)", NAVY, sz=11)
kpis = [
    ("KPI-01", "Intensidad por hectárea productiva", f"=G{r_net}/900", "tCO2e/ha"),
    ("KPI-02", "Intensidad de emisiones Scope 1", f"=SUM(G6:G11)/G{r_sub_em}*100", "% sobre total"),
    ("KPI-03", "Intensidad de emisiones Scope 2", f"=G12/G{r_sub_em}*100", "% sobre total"),
    ("KPI-04", "Proporción capturada por sumideros", f"=IFERROR(ABS(G{r_sub_sk})/G{r_sub_em}*100,0)", "% capturas/emisiones"),
    ("KPI-05", "Huella por cabeza ganadera (ingresar total cabezas)", "=IFERROR(G21/D28,0)", "tCO2e/cabeza/año"),
    ("KPI-06", "Huella por tonelada de grano producida (ingresar prod.)", "=IFERROR(G21/E28,0)", "tCO2e/t grano"),
    ("KPI-07", "Captura neta de sumideros (valor absoluto)", f"=ABS(G{r_sub_sk})", "tCO2e capturas/año"),
    ("KPI-08", "Balance carbono (positivo = emisor neto)", f"=G{r_net}", "tCO2e neto/año"),
]
for j, text in enumerate(["Código", "Indicador", "Fórmula / Valor [CALC]", "Unidad", "Valor ingresado (si aplica)", "", "", "", "", "Notas"]):
    header(ws10, 24, j + 1, text, bg=NAVY, sz=9)

for i, (code, name, formula, unit) in enumerate(kpis):
    r = 25 + i
    ws10.row_dimensions[r].height = 20
    bg_r = OR_L if i % 2 == 0 else 'FFFFFF'
    set_cell(ws10, r, 1, code, bg=bg_r, bold=True, halign='center')
    ws10.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2)
    lbl(ws10, r, 2, name, bg_r)
    calc(ws10, r, 3, formula)
    set_cell(ws10, r, 4, unit, bg=bg_r, halign='center')
    inp(ws10, r, 5, None)

# Datos auxiliares para KPIs (cabezas, producción)
r_aux = 33
merge(ws10, r_aux, 1, r_aux, 10, "DATOS AUXILIARES PARA KPIs (ingresar al final del período)", GR2, '000000', bold=True, sz=10)
aux_data = [
    ("Total de cabezas de ganado (promedio anual)", None, "cabezas", "Para KPI-05"),
    ("Producción total de granos (t/año)", None, "toneladas", "Para KPI-06"),
    ("Litros de leche producidos (L/año)", None, "litros/año", "Para tambo"),
    ("Kg de carne producidos (kg res/año)", None, "kg res/año", "Para feedlot"),
    ("Ingresos totales del establecimiento ($/año)", None, "$", "Opcional: para huella por $ facturado"),
]
for i, (field, val, unit, note) in enumerate(aux_data):
    r = r_aux + 1 + i
    ws10.row_dimensions[r].height = 20
    lbl(ws10, r, 1, field, GR1, bold=True)
    ws10.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    inp(ws10, r, 2, val)
    lbl(ws10, r, 4, unit, GR1)
    ws10.merge_cells(start_row=r, start_column=5, end_row=r, end_column=10)
    lbl(ws10, r, 5, note, GR1)

set_col_widths(ws10, [8, 38, 12, 12, 12, 12, 14, 12, 18, 22])

# ─────────────────────────────────────────────────────────────────────────────
# GUARDAR
# ─────────────────────────────────────────────────────────────────────────────
output_path = "/Users/martinhulais/Downloads/Matriz_HuellaCarbono_Agropecuario.xlsx"
wb.save(output_path)
print(f"Archivo guardado: {output_path}")
print(f"Hojas: {[ws.title for ws in wb.worksheets]}")
